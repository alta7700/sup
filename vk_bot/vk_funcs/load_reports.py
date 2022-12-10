from ReportsDjango import settings
from .config import *
import locale
from .all_buttons import *
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from .excels_styles import *
from .instructions import right_attandance_tags

locale.setlocale(locale.LC_ALL, 'ru_RU.utf8')


# Добавление рапортов
########
def send_report_message_pattern(user_id):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return
    if not (head_info['is_head'] or head_info['is_head_assistant']):
        var_message(user_id, 'Ты не староста, откуда кнопка😕', kb_params=main_kb_student)
        return

    faculty = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    cur.execute(
        '''SELECT min(pr_lk_datetime) as min_date FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s''',
        (faculty, course_n, group_n, False)
    )
    min_date = cur.fetchone()

    if not min_date['min_date']:
        var_message(user_id, 'Игра пройдена, поздравляю, все рапорты отправлены', kb_params=main_kb_head)
        return
    min_date = min_date['min_date'].replace(tzinfo=None)

    if not min_date < datetime.now():
        next_p = f'Ближайшая пара {min_date.strftime("%d.%m в %H:%M")}'
        var_message(user_id, f'На данный момент все рапорты отправлены\n{next_p}', kb_params=main_kb_head)
        return
    cur.execute(
        '''SELECT * FROM reports_pairsschedule WHERE
        faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s AND pr_lk_datetime = %s''',
        (faculty, course_n, group_n, False, min_date)
    )
    pair_info = cur.fetchone()
    cur.execute(
        '''SELECT * FROM reports_student
        WHERE faculty_id = %s AND course_n = %s and group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty, course_n, group_n, False)
    )
    group_stud = cur.fetchall()
    subject = get_subject_by_id(pair_info["subject_id"])
    if pair_info['is_practical']:
        pattern = f'Рапорт\nПрактическое занятие\n{subject["short_title"]}\n{min_date.strftime("%d.%m.%Y %H:%M")}'
    else:
        pattern = f'Рапорт\nЛекция\n{subject["short_title"]}\n{min_date.strftime("%d.%m.%Y %H:%M")}'
    for stud in group_stud:
        pattern += f'\n{stud["position_in_group"]}. {stud["surname"]}:'
    var_message(user_id, pattern, kb_params=main_kb_head)


def parse_report_message(user_id, msg):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return
    if not (head_info['is_head'] or head_info['is_head_assistant']):
        var_message(user_id, 'Ты не староста и даже не его помощник', kb_params=main_kb_student)
        return

    faculty = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    cur.execute(
        '''SELECT min(pr_lk_datetime) as min_date FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s''',
        (faculty, course_n, group_n, False)
    )
    min_date = cur.fetchone()

    if not min_date['min_date']:
        var_message(user_id, 'Незаполненных рапортов нет', kb_params=main_kb_head)
        return
    min_date = min_date['min_date'].replace(tzinfo=None)

    if not min_date < datetime.now():
        next_p = f'Этот рапорт не зачту\nБлижайшая пара {min_date.strftime("%d.%m в %H:%M")}'
        var_message(user_id, f'На данный момент все рапорты отправлены\n{next_p}', kb_params=main_kb_head)
        return

    msg = [m.strip() for m in msg.split('\n')]
    if len(msg) <= 4:
        var_message(user_id, 'Чтобы получить шаблон, напиши:\nРапорт таблицей\nили\nРапорт сообщением')
        return
    cur.execute(
        '''SELECT * FROM reports_pairsschedule WHERE
        faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s AND pr_lk_datetime = %s''',
        (faculty, course_n, group_n, False, min_date)
    )
    pair_info = cur.fetchone()
    subject = get_subject_by_id(pair_info["subject_id"])
    if pair_info['is_practical']:
        pattern = f'Рапорт\nПрактическое занятие\n{subject["short_title"]}\n{min_date.strftime("%d.%m.%Y %H:%M")}'
    else:
        pattern = f'Рапорт\nЛекция\n{subject["short_title"]}\n{min_date.strftime("%d.%m.%Y %H:%M")}'
    if '\n'.join(msg[:4]) != pattern:
        var_message(user_id, 'Ничего не меняй в шаблоне, ТОЛЬКО проставь посещаемость')
        return

    cur.execute(
        '''SELECT * FROM reports_student
        WHERE faculty_id = %s AND course_n = %s and group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty, course_n, group_n, False)
    )
    group_stud = cur.fetchall()
    num = 4
    if len(group_stud) == len(msg[4:]):
        missings = []
        for stud in group_stud:
            if f'{stud["position_in_group"]}. {stud["surname"]}' == msg[num].split(':')[0]:
                try:
                    reason = msg[num].split(':')[1].strip()
                except IndexError:
                    var_message(user_id, 'Нужно проставить посещаемость всем студентам')
                    return
                if reason not in ('+', 'н', 'Н', 'б', 'Б', 'с', 'С', 'c', 'C'):
                    var_message(user_id, right_attandance_tags)
                    return
                if reason in ('н', 'Н'):
                    missings.append((stud['id'], pair_info['id'], 'н', True))
                elif reason in ('б', 'Б'):
                    missings.append((stud['id'], pair_info['id'], 'б', True))
                elif reason in ('c', 'С', 'с', 'С'):
                    missings.append((stud['id'], pair_info['id'], 'с', False))
                elif reason == '+':
                    pass
            else:
                var_message(user_id, 'Ничего не меняй в шаблоне, ТОЛЬКО проставь посещаемость')
                return
            num += 1
        if missings:
            cur.executemany(
                '''INSERT INTO reports_missingpair(student_id, pair_id, missing_reason, is_debt)
                VALUES (%s, %s, %s, %s)''',
                missings
            )
        cur.execute(
            'UPDATE reports_pairsschedule SET is_reported = %s WHERE id = %s',
            (True, pair_info['id'])
        )
        # connection.commit()
        var_message(user_id, 'Рапорт принят')
    else:
        var_message(user_id, 'Ничего не меняй в шаблоне, ТОЛЬКО проставь посещаемость')
        return


def send_report_excel_pattern(user_id):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return
    if not (head_info['is_head'] or head_info['is_head_assistant']):
        var_message(user_id, 'Ты не староста, откуда кнопка😕', kb_params=main_kb_student)
        return
    faculty = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    cur.execute(
        '''SELECT min(pr_lk_datetime) as min_date FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s''',
        (faculty, course_n, group_n, False)
    )
    min_date = cur.fetchone()

    if not min_date['min_date']:
        var_message(user_id, 'Незаполненных рапортов нет', kb_params=main_kb_head)
        return
    min_date = min_date['min_date'].replace(tzinfo=None)

    if min_date > datetime.now():
        next_p = f'Ближайшая пара {min_date.strftime("%d.%m в %H:%M")}'
        var_message(user_id, f'На данный момент все рапорты отправлены\n{next_p}', kb_params=main_kb_head)
        return

    end_of_min_date_week = (min_date + timedelta(days=6-min_date.weekday())).replace(hour=23, minute=59)
    if end_of_min_date_week < datetime.now():
        end_date = end_of_min_date_week
    else:
        end_date = datetime.now()
    cur.execute(
        '''SELECT * FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s
        AND pr_lk_datetime BETWEEN %s AND %s
        ORDER BY pr_lk_datetime''',
        (faculty, course_n, group_n, False, min_date, end_date)
    )
    pairs_info = cur.fetchall()
    cur.execute(
        '''SELECT * FROM reports_student
        WHERE faculty_id = %s AND course_n = %s and group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty, course_n, group_n, False)
    )
    group_stud = cur.fetchall()

    wb = Workbook()
    wb_sheet = wb.active
    wb_sheet.cell(row=1, column=1).value = '№'
    wb_sheet.cell(row=1, column=1).border = thin_border
    wb_sheet.merge_cells(start_row=1, end_row=5, start_column=1, end_column=1)
    wb_sheet.column_dimensions['A'].width = 4
    wb_sheet.cell(row=1, column=2).value = 'Фамилия Имя'
    wb_sheet.cell(row=1, column=2).border = thin_border
    wb_sheet.merge_cells(start_row=1, end_row=5, start_column=2, end_column=2)
    stud_names_len = []
    for i, stud in enumerate(group_stud):
        wb_sheet.cell(row=i+6, column=1).value = stud['position_in_group']
        wb_sheet.cell(row=i+6, column=1).border = thin_border
        stud_name = f'{stud["surname"]} {stud["name"]}'
        wb_sheet.cell(row=i+6, column=2).value = stud_name
        wb_sheet.cell(row=i+6, column=2).border = thin_border
        stud_names_len.append(len(stud_name))
    wb_sheet.column_dimensions['B'].width = max(stud_names_len) * 1.2
    start_this_date_col = 3
    extra_cols = 0
    this_date = min_date.date()
    for i, pair_info in enumerate(pairs_info):
        pair_date = pair_info['pr_lk_datetime'].date()
        if this_date != pair_date:
            wb_sheet.cell(row=1, column=start_this_date_col).value = this_date.strftime('%A\n%d.%m.%Y')
            wb_sheet.cell(row=1, column=start_this_date_col).border = thin_border
            wb_sheet.cell(row=1, column=start_this_date_col).alignment = date_alignment
            wb_sheet.merge_cells(start_row=1, start_column=start_this_date_col,
                                 end_row=2, end_column=start_this_date_col + 3 + extra_cols)
            start_this_date_col += 4 + extra_cols
            extra_cols = 0
            this_date = pair_date
        if i + 1 == len(pairs_info):
            wb_sheet.cell(row=1, column=start_this_date_col).value = this_date.strftime('%A\n%d.%m.%Y')
            wb_sheet.cell(row=1, column=start_this_date_col).border = thin_border
            wb_sheet.cell(row=1, column=start_this_date_col).alignment = date_alignment
            wb_sheet.merge_cells(start_row=1, start_column=start_this_date_col,
                                 end_row=2, end_column=start_this_date_col + 3 + extra_cols)
        if pair_info['pair_n'] > 4:
            extra_cols = max(extra_cols, pair_info['pair_n'] - 4)
        this_col = pair_info['pair_n'] + start_this_date_col - 1
        wb_sheet.cell(row=3, column=this_col).value = pair_info['pair_n']
        wb_sheet.cell(row=3, column=this_col).alignment = Alignment(horizontal='center')
        wb_sheet.cell(row=3, column=this_col).border = thin_border

        wb_sheet.cell(row=4, column=this_col).value = 'пр' if pair_info['is_practical'] else 'лк'
        wb_sheet.cell(row=4, column=this_col).alignment = Alignment(horizontal='center')
        wb_sheet.cell(row=4, column=this_col).border = thin_border

        wb_sheet.cell(row=5, column=this_col).value = get_subject_by_pair(pair_info)['short_title']
        wb_sheet.cell(row=5, column=this_col).alignment = Alignment(horizontal='center', textRotation=90)
        wb_sheet.cell(row=5, column=this_col).border = thin_border
    for col in range(3, wb_sheet.max_column + 1):
        wb_sheet.column_dimensions[wb_sheet.cell(row=3, column=col).column_letter].width = 4
        start_row = 3 if not wb_sheet.cell(row=3, column=col).value else 6
        for r in range(start_row, wb_sheet.max_row + 1):
            wb_sheet.cell(row=r, column=col).alignment = Alignment(horizontal='center')
            wb_sheet.cell(row=r, column=col).border = thin_border

    min_d = min_date.strftime('%d.%m.%Y')
    max_d = end_date.strftime('%d.%m.%Y')
    wb_title = f'{get_faculty_by_id(faculty)["short_title"]} {course_n}курс {group_n}гр рапорт {min_d}-{max_d}.xlsx'
    wb.save(wb_title)
    doc_sender(user_id, wb_title, 'Заполни этот рапорт и отправь файл вместе с сообщением "Рапорт"',
               kb_params=main_kb_head)


def parse_report_excel_pattern(user_id, attach):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return
    if not (head_info['is_head'] or head_info['is_head_assistant']):
        var_message(user_id, 'Ты не староста и даже не его помощник', kb_params=main_kb_student)
        return
    if not download_attach_xlsx(user_id, attach):
        os.remove(attach['title'])
        return

    faculty = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    cur.execute(
        '''SELECT min(pr_lk_datetime) as min_date FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s''',
        (faculty, course_n, group_n, False)
    )
    min_date = cur.fetchone()
    if not min_date['min_date']:
        var_message(user_id, 'Незаполненных рапортов нет, так что этот не приму')
        return
    min_date = min_date['min_date'].replace(tzinfo=None)

    if min_date > datetime.now():
        next_p = f'Этот рапорт не зачту\nБлижайшая пара {min_date.strftime("%d.%m в %H:%M")}'
        var_message(user_id, f'На данный момент все рапорты отправлены\n{next_p}')
        return

    end_of_min_date_week = (min_date + timedelta(days=6-min_date.weekday())).replace(hour=23, minute=59)
    if end_of_min_date_week < datetime.now():
        end_date = end_of_min_date_week
    else:
        end_date = datetime.now()
    cur.execute(
        '''SELECT * FROM reports_pairsschedule WHERE
        faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s AND pr_lk_datetime BETWEEN %s AND %s
        ORDER BY pr_lk_datetime''',
        (faculty, course_n, group_n, False, min_date, end_date)
    )
    pairs_info = cur.fetchall()
    cur.execute(
        '''SELECT * FROM reports_student WHERE faculty_id = %s AND course_n = %s and group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty, course_n, group_n, False)
    )
    group_stud = cur.fetchall()

    # [0][0] и [0] пустые чтобы индексы были равны с excel
    wb_matrix = [[], [None, '№', 'Фамилия Имя', None, None, None, None]]
    for _ in range(4):
        wb_matrix.append([None, None, None, None, None, None, None])
    for stud in group_stud:
        wb_matrix.append([None, stud['position_in_group'], f'{stud["surname"]} {stud["name"]}', None, None, None, None])

    start_this_date_col = 3
    extra_cols = 0
    this_date = min_date.date()
    for i, pair_info in enumerate(pairs_info):
        pair_date = pair_info['pr_lk_datetime'].date()
        if this_date != pair_date:
            wb_matrix[1][start_this_date_col] = this_date.strftime('%A\n%d.%m.%Y')
            for index in range(len(wb_matrix)):
                wb_matrix[index] = wb_matrix[index] + [None, None, None, None]
            start_this_date_col += 4 + extra_cols
            extra_cols = 0
            this_date = pair_date
        if i + 1 == len(pairs_info):
            wb_matrix[1][start_this_date_col] = this_date.strftime('%A\n%d.%m.%Y')

        if pair_info['pair_n'] > 4:
            if extra_cols < max(extra_cols, pair_info['pair_n'] - 4):
                new_cols = pair_info['pair_n'] - 4 - extra_cols
                extra_cols = max(extra_cols, pair_info['pair_n'] - 4)
                for index in range(len(wb_matrix)):
                    wb_matrix[index] = wb_matrix[index] + [None for _ in new_cols]
        this_col = pair_info['pair_n'] + start_this_date_col - 1
        wb_matrix[3][this_col] = pair_info['pair_n']
        wb_matrix[4][this_col] = 'пр' if pair_info['is_practical'] else 'лк'
        wb_matrix[5][this_col] = get_subject_by_pair(pair_info)['short_title']

    book = load_workbook(attach['title'])
    sheet = book.active
    is_everything_ok = True
    difference_list = []
    try:
        for row in list(sheet.rows)[:5]:
            for cell in row[2:]:
                if cell.value != wb_matrix[cell.row][cell.column]:
                    is_everything_ok = False
                    difference_list.append(f'{wb_matrix[cell.row][cell.column]} на {cell.value}\n')
        for col in list(sheet.columns)[:2]:
            for cell in col:
                if cell.value != wb_matrix[cell.row][cell.column]:
                    is_everything_ok = False
                    difference_list.append(f'{wb_matrix[cell.row][cell.column]} на {cell.value}\n')
    except IndexError:
        is_everything_ok = False
        difference_list = [1, 1, 1, 1, 1, 1]  # типа больше 5
    if not is_everything_ok:
        if len(difference_list) <= 5:
            var_message(user_id, f'Рапорт не принят\nПохоже ты заменил:\n{"".join(difference_list)}')
        else:
            var_message(user_id, 'Рапорт не принят\nМного изменений в шаблоне')
        return
    attendance_dict = {}
    this_date = min_date.date()
    for col in list(sheet.columns)[2:]:
        if col[0].value:
            this_date = datetime.strptime(col[0].value, '%A\n%d.%m.%Y').date()
            attendance_dict[this_date] = {}
        if col[2].value:
            pair_num = str(col[2].value)
            attendance = {}
            for cell in col[5:]:
                if type(cell.value) == str:
                    cell_value = cell.value.strip()
                else:
                    var_message(user_id, 'Что-то не так с метками посещаемости')
                    return
                if cell_value:
                    if cell_value not in ('+', 'н', 'Н', 'б', 'Б', 'с', 'С', 'c', 'C'):
                        var_message(user_id, right_attandance_tags)
                        return
                    position = str(sheet.cell(row=cell.row, column=1).value)
                    if cell_value in ('н', 'Н'):
                        attendance[position] = 'н'
                    elif cell_value in ('б', 'Б'):
                        attendance[position] = 'б'
                    elif cell_value in ('c', 'С', 'с', 'C'):
                        attendance[position] = 'с'
                    elif cell_value == '+':
                        attendance[position] = '+'
                else:
                    var_message(user_id, 'Нужно проставить посещаемость всем студентам')
                    return
            attendance_dict[this_date][pair_num] = {
                'pr_or_lk': col[3].value.strip(),
                'subject': col[4].value.strip(),
                'attendance': attendance
            }
    missings = []
    for pair_info in pairs_info:
        pair_date = pair_info['pr_lk_datetime'].date()
        if attendance_dict.get(pair_date):
            pair_n = str(pair_info['pair_n'])
            if attendance_dict[pair_date].get(pair_n):
                cur.execute('UPDATE reports_pairsschedule SET is_reported = %s WHERE id = %s', (True, pair_info['id']))
                pair_n_missings = attendance_dict[pair_date][pair_n]['attendance']
                for stud in group_stud:
                    position = str(stud['position_in_group'])
                    if pair_n_missings.get(position):
                        reason = pair_n_missings[position]
                        if reason != '+':
                            if reason == 'н':
                                missings.append((stud['id'], pair_info['id'], reason, True))
                            elif reason == 'б':
                                missings.append((stud['id'], pair_info['id'], reason, True))
                            elif reason == 'с':
                                missings.append((stud['id'], pair_info['id'], reason, False))
    if missings:
        cur.executemany(
            '''INSERT INTO reports_missingpair(student_id, pair_id, missing_reason, is_debt)
            VALUES (%s, %s, %s, %s)''',
            missings
        )
    # connection.commit()
    var_message(user_id, 'Рапорт принят')
    os.remove(attach['title'])
########


# Изменение рапортов
########
def show_subjects_to_change_report(user_id, btn_payload: str, is_practical: bool):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return

    if not (head_info['is_head'] or head_info['is_head_assistant']):
        var_message(user_id, 'Ты не староста, откуда кнопка😕', kb_params=main_kb_student)
        return

    faculty_id = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    now_dt = datetime.now()
    cur.execute(
        '''SELECT
            s.id AS id,
            s.short_title AS short_title,
            MIN(ps.pr_lk_datetime) AS min_date
        FROM
            reports_pairsschedule AS ps
            JOIN reports_subject AS s
            ON s.id = ps.subject_id
        WHERE
            ps.faculty_id = %s AND
            ps.course_n = %s AND
            ps.group_n = %s AND
            ps.is_reported = %s AND
            ps.is_practical = %s AND
            ps.pr_lk_datetime BETWEEN %s AND %s
        GROUP BY s.id, s.short_title
        ORDER BY min_date''',
        (faculty_id, course_n, group_n, True, is_practical, now_dt-timedelta(days=settings.CHANGE_REPORT_DAYS), now_dt)
    )
    subjects = cur.fetchall()
    if subjects:
        num = 1
        subjects_kb = {num: []}
        max_btn_in_line = 3 if len(subjects) > 16 else 2
        for subject in subjects:
            subjects_kb[num].append((subject['short_title'], 'primary', btn_payload.format(subject["id"])))
            if len(subjects_kb[num]) == max_btn_in_line:
                num += 1
                subjects_kb[num] = []
        subjects_kb[num].append(cancel_btn)
        var_message(user_id, '...', kb_params=subjects_kb)
    else:
        if is_practical:
            var_message(user_id, 'Ни по одному практическому занятию за последние 2 недели ещё нет рапорта',
                        kb_params=main_kb_head)
        else:
            var_message(user_id, 'Ни по одной лекции за последние 2 недели ещё нет рапорта', kb_params=main_kb_head)


def show_subject_dates_to_change_report(user_id, btn_payload: str, is_practical: bool):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return

    if not (head_info['is_head'] or head_info['is_head_assistant']):
        var_message(user_id, 'Ты не староста, откуда кнопка😕', kb_params=main_kb_student)
        return
    faculty_id = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    subject_id = int(btn_payload[:-2].split('-')[1])
    now_dt = datetime.now()
    cur.execute(
        '''SELECT
            id,
            pr_lk_datetime
        FROM
            reports_pairsschedule
        WHERE
            faculty_id = %s AND
            course_n = %s AND
            group_n = %s AND
            subject_id = %s AND 
            is_reported = %s AND
            is_practical = %s AND
            pr_lk_datetime BETWEEN %s AND %s
        ORDER BY pr_lk_datetime''',
        (faculty_id, course_n, group_n, subject_id, True, is_practical,
         now_dt-timedelta(days=settings.CHANGE_REPORT_DAYS), now_dt)
    )
    subject_datetimes = cur.fetchall()
    if subject_datetimes:
        btn_payload = f'{{{btn_payload[:-2]}-{{}}"}}}}'
        num = 1
        dates_kb = {num: []}
        date_btn_line = []
        for subj_dt in subject_datetimes:
            dt = subj_dt['pr_lk_datetime'].strftime('%d.%m.%Y')
            if len(date_btn_line) == 3:
                dates_kb[num] = date_btn_line
                date_btn_line = []
                num += 1
            date_btn_line.append((dt, 'primary', btn_payload.format(subj_dt['id'])))
        if len(date_btn_line) != 3:
            date_btn_line.append(cancel_btn)
            dates_kb[num] = date_btn_line
        else:
            dates_kb[num] = date_btn_line
            dates_kb[num + 1] = [cancel_btn, ]
        var_message(user_id, '...', kb_params=dates_kb)
    else:
        subject = get_subject_by_id(subject_id)
        if subject:
            s_title = subject['short_title']
            var_message(user_id, f'Нет ни одного рапорта за {"практические занятия" if is_practical else "лекции"} по '
                                 f'{s_title}, такого быть не должно, обратись к [id39398636|Саше]',
                        kb_params=main_kb_head)
        else:
            var_message(user_id, 'Что-то не так, не могу найти этот предмет\nПопробуй выйти в начало и нажми все те же '
                                 'кнопки заново\nЕсли не получится, обратись к [id39398636|Саше]')
        var_message(672645458, f'ID предмета - {subject_id}\nvk_id - {user_id}\npayload - {btn_payload}')


def send_report_message_pattern_to_change_raport(user_id, btn_payload):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
    if not (head_info['is_head'] or head_info['is_head_assistant']):
        var_message(user_id, 'Ты не староста, откуда кнопка😕', kb_params=main_kb_student)
        return

    faculty_id = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    pair_id = int(btn_payload[:-2].split('-')[2])
    cur.execute(
        '''SELECT
            ps.is_practical AS is_practical,
            s.short_title AS subject,
            ps.pr_lk_datetime AS pr_lk_datetime,
            ps.is_reported AS is_reported
        FROM reports_pairsschedule as ps
            JOIN reports_subject as s
            ON s.id = subject_id
        WHERE
            ps.id = %s''',
        (pair_id, )
    )
    pair_info = cur.fetchone()
    pair_datetime = pair_info['pr_lk_datetime'].replace(tzinfo=None)
    if (datetime.now() - pair_datetime).days > settings.CHANGE_REPORT_DAYS:
        var_message(user_id, f'Прошло больше {settings.CHANGE_REPORT_DAYS} дней, изменять нельзя')
        return
    cur.execute(
        '''SELECT
            st.pos AS pos,
            st.surname AS surname,
            mp.missing_reason AS reason
        FROM (SELECT * FROM reports_missingpair WHERE pair_id = %s) AS mp
            RIGHT JOIN (
            SELECT 
                id AS id,
                surname AS surname,
                position_in_group AS pos
            FROM reports_student AS st
            WHERE
                faculty_id = %s AND
                course_n = %s AND
                group_n = %s AND
                is_fired = %s) AS st
            ON st.id = mp.student_id
        ORDER BY st.id''',
        (pair_id, faculty_id, course_n, group_n, False)
    )
    group_stud = cur.fetchall()
    pair_date = pair_datetime.strftime('%d.%m.%Y %H:%M')
    if pair_info['is_practical']:
        pattern = f'Рапорт ИЗМЕНИТЬ\nПрактическое занятие\n{pair_info["subject"]}\n{pair_date}'
    else:
        pattern = f'Рапорт ИЗМЕНИТЬ\nЛекция\n{pair_info["subject"]}\n{pair_date}'
    for stud in group_stud:
        pattern += f'\n{stud["pos"]}. {stud["surname"]}:{stud["reason"] if stud["reason"] else "+"}'
    var_message(user_id, pattern, kb_params=main_kb_head)


def parse_report_change_message(user_id, msg):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return
    if not (head_info['is_head'] or head_info['is_head_assistant']):
        var_message(user_id, 'Ты не староста и даже не его помощник', kb_params=main_kb_student)
        return

    faculty_id = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    msg = [x.strip() for x in msg.split('\n')[1:]]
    if len(msg) < 4:
        var_message(user_id, 'Ничего не меняй в шаблоне, ТОЛЬКО проставь посещаемость')
        return
    if msg[0] == 'Практическое занятие':
        is_practical = True
    elif msg[0] == 'Лекция':
        is_practical = False
    else:
        var_message(user_id, 'Ничего не меняй в шаблоне, ТОЛЬКО проставь посещаемость')
        return
    cur.execute(
        '''SELECT * FROM reports_subject WHERE
        faculty_id = %s AND course_n = %s AND short_title = %s AND half = %s''',
        (faculty_id, course_n, msg[1], 1)
    )
    subject = cur.fetchone()
    if not subject:
        var_message(user_id, f'Предмета {msg[1]} нет, ничего не меняй в шаблоне, ТОЛЬКО проставь посещаемость')
        return
    try:
        pair_datetime = datetime.strptime(msg[2], '%d.%m.%Y %H:%M')
    except ValueError:
        var_message(user_id, 'Не трогай дату и время, ТОЛЬКО проставь посещаемость')
        return
    cur.execute(
        '''SELECT * FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND
        is_practical = %s AND subject_id = %s AND pr_lk_datetime = %s''',
        (faculty_id, course_n, group_n, is_practical, subject['id'], pair_datetime)
    )
    pair_info = cur.fetchone()
    if not pair_info:
        var_message(user_id, 'Такой пары нет')
        return
    if not pair_info['is_reported']:
        var_message(user_id, 'Рапорт для этой пары ещё не отправлен')
        return
    if (datetime.now() - pair_datetime).days > settings.CHANGE_REPORT_DAYS:
        var_message(user_id, f'Прошло больше {settings.CHANGE_REPORT_DAYS} дней, изменять нельзя')
        return
    cur.execute(
        '''SELECT id, position_in_group as pos, surname FROM reports_student
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty_id, course_n, group_n, False)
    )
    group_stud = cur.fetchall()
    if len(group_stud) > len(msg[3:]):
        var_message(user_id, 'Тут кого-то не хватает')
        return
    elif len(group_stud) < len(msg[3:]):
        var_message(user_id, 'Тут кто-то лишний')
        return

    missings = []
    for i, stud in enumerate(group_stud):
        if f'{stud["pos"]}. {stud["surname"]}' == msg[i+3].split(':')[0].strip():
            try:
                reason = msg[i+3].split(':')[1].strip()
            except IndexError:
                var_message(user_id, 'Нужно проставить посещаемость всем студентам')
                return
            if reason not in ('+', 'н', 'Н', 'б', 'Б', 'с', 'С', 'c', 'C'):
                var_message(user_id, right_attandance_tags)
                return
            if reason in ('н', 'Н'):
                missings.append((stud['id'], pair_info['id'], 'н', True))
            elif reason in ('б', 'Б'):
                missings.append((stud['id'], pair_info['id'], 'б', True))
            elif reason in ('c', 'С', 'с', 'C'):
                missings.append((stud['id'], pair_info['id'], 'с', False))
            elif reason == '+':
                pass
        else:
            var_message(user_id, f'Не меняй фамилии студентов и их номера')
            return
    cur.execute('SELECT * FROM reports_missingpair WHERE pair_id = %s ORDER BY student_id', (pair_info['id'], ))
    missings_db = [(x['student_id'], x['pair_id'], x['missing_reason']) for x in cur.fetchall()]
    if not missings_db and not missings:
        var_message(user_id, 'Нет изменений')
        return

    attendance_change = []
    insert_missings = []
    missings_db_dict = {x[0]: x for x in missings_db}
    missings_dict = {x[0]: (x[0], x[1], x[2], x[3]) for x in missings}
    time_now = datetime.now().strftime('%Y.%m.%d %H:%M:%S')
    if missings_db:
        if missings_db == [(x[0], x[1], x[2]) for x in missings]:
            var_message(user_id, 'Нет изменений')
            return
        for stud in group_stud:
            if missings_db_dict.get(stud['id']):
                reason_db = missings_db_dict[stud['id']][2]
                if missings_dict.get(stud['id']):
                    new_reason = missings_dict[stud['id']][2]
                    if reason_db != new_reason:
                        attendance_change.append((reason_db, new_reason, time_now, pair_info['id'], stud['id']))
                        insert_missings.append(missings_dict[stud['id']])
                else:
                    attendance_change.append((reason_db, "+", time_now, pair_info['id'], stud['id']))
            elif missings_dict.get(stud['id']):
                new_reason = missings_dict[stud['id']][2]
                attendance_change.append(('+', new_reason, time_now, pair_info['id'], stud['id']))
                insert_missings.append(missings_dict[stud['id']])

    elif missings:
        if not missings_db:
            insert_missings = missings
            for m in missings:
                attendance_change.append(('+', m[2], time_now, m[1], m[0]))
    else:
        var_message(user_id, 'Нет изменений')
        return

    if attendance_change:
        cur.executemany(
            '''INSERT INTO
            reports_changesattendance(past_missing_reason, new_missing_reason, change_datetime, pair_id, student_id)
            VALUES (%s, %s, %s, %s, %s)
            ''',
            attendance_change
        )
    else:
        var_message(user_id, 'Нет изменений')
        return

    if missings_db:
        cur.execute(
            'DELETE FROM reports_missingpair WHERE pair_id = %s AND student_id = ANY(%s)',
            (pair_info['id'], list((x[4] for x in attendance_change))))
    if missings:
        cur.executemany(
            'INSERT INTO reports_missingpair(student_id, pair_id, missing_reason, is_debt) VALUES (%s, %s, %s, %s)',
            insert_missings
        )

    var_message(user_id, 'Рапорт изменен')
########
