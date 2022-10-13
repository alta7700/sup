import locale
from ..config import *
from ..excels_styles import *
from random import choice
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook


locale.setlocale(locale.LC_ALL, ('ru_RU', 'UTF-8'))


def test_send_report_message_pattern(user_id, end_date):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        # var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return 1
    if not (head_info['is_head'] or head_info['is_head_assistant']):
        # var_message(user_id, 'Ты не староста, откуда кнопка😕', kb_params=main_kb_student)
        return 2

    faculty = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    cur.execute(
        '''SELECT min(pr_lk_datetime) as min_date FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s''',
        (faculty, course_n, group_n, False)
    )
    min_date = cur.fetchone()

    if not min_date['min_date']:
        # var_message(user_id, 'Игра пройдена, поздравляю, все рапорты отправлены', kb_params=main_kb_head)
        return 3
    min_date = min_date['min_date'].replace(tzinfo=None)

    if not min_date < end_date:
        # next_p = f'Ближайшая пара {min_date.strftime("%d.%m в %H:%M")}'
        # var_message(user_id, f'На данный момент все рапорты отправлены\n{next_p}', kb_params=main_kb_head)
        return 4
    cur.execute(
        '''SELECT * FROM reports_pairsschedule WHERE
        faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s AND pr_lk_datetime = %s''',
        (faculty, course_n, group_n, False, min_date)
    )
    pair_info = cur.fetchone()
    cur.execute(
        '''SELECT * FROM reports_student
        WHERE faculty_id = %s AND course_n = %s and group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty, course_n, group_n, False)
    )
    group_stud = cur.fetchall()
    subject = get_subject_by_id(pair_info["subject_id"])
    if pair_info['is_practical']:
        pattern = f'Рапорт\nПрактическое занятие\n{subject["short_title"]}\n{min_date.strftime("%d.%m.%Y %H:%M")}'
    else:
        pattern = f'Рапорт\nЛекция\n{subject["short_title"]}\n{min_date.strftime("%d.%m.%Y %H:%M")}'
    for stud in group_stud:
        pattern += f'\n{stud["position_in_group"]}. {stud["surname"]}:{choice("++++++++нбс")}'
    # var_message(user_id, pattern, kb_params=main_kb_head)
    test_parse_report_message(user_id, pattern, end_date)


def test_parse_report_message(user_id, msg, end_date):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        # var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return 5
    if not (head_info['is_head'] or head_info['is_head_assistant']):
        # var_message(user_id, 'Ты не староста и даже не его помощник', kb_params=main_kb_student)
        return 6

    faculty = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    cur.execute(
        '''SELECT min(pr_lk_datetime) as min_date FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s''',
        (faculty, course_n, group_n, False)
    )
    min_date = cur.fetchone()

    if not min_date['min_date']:
        # var_message(user_id, 'Незаполненных рапортов нет', kb_params=main_kb_head)
        return 7
    min_date = min_date['min_date'].replace(tzinfo=None)

    if not min_date < end_date:
        # next_p = f'Этот рапорт не зачту\nБлижайшая пара {min_date.strftime("%d.%m в %H:%M")}'
        # var_message(user_id, f'На данный момент все рапорты отправлены\n{next_p}', kb_params=main_kb_head)
        return 8

    msg = [m.strip() for m in msg.split('\n')]
    if len(msg) <= 4:
        # var_message(user_id, 'Чтобы получить шаблон, напиши:\nРапорт таблицей\nили\nРапорт сообщением')
        return 9
    cur.execute(
        '''SELECT * FROM reports_pairsschedule WHERE
        faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s AND pr_lk_datetime = %s''',
        (faculty, course_n, group_n, False, min_date)
    )
    pair_info = cur.fetchone()
    subject = get_subject_by_id(pair_info["subject_id"])
    if pair_info['is_practical']:
        pattern = f'Рапорт\nПрактическое занятие\n{subject["short_title"]}\n{min_date.strftime("%d.%m.%Y %H:%M")}'
    else:
        pattern = f'Рапорт\nЛекция\n{subject["short_title"]}\n{min_date.strftime("%d.%m.%Y %H:%M")}'
    if '\n'.join(msg[:4]) != pattern:
        # var_message(user_id, 'Ничего не меняй в шаблоне, ТОЛЬКО проставь посещаемость')
        return 10

    cur.execute(
        '''SELECT * FROM reports_student
        WHERE faculty_id = %s AND course_n = %s and group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty, course_n, group_n, False)
    )
    group_stud = cur.fetchall()
    num = 4
    if len(group_stud) == len(msg[4:]):
        missings = []
        for stud in group_stud:
            if f'{stud["position_in_group"]}. {stud["surname"]}' == msg[num].split(':')[0]:
                try:
                    reason = msg[num].split(':')[1].strip()
                except IndexError:
                    # var_message(user_id, 'Нужно проставить посещаемость всем студентам')
                    return 11
                if reason not in ('+', 'н', 'Н', 'б', 'Б', 'с', 'С', 'c', 'C'):
                    # var_message(user_id, right_attandance_tags)
                    return 12
                if reason in ('н', 'Н'):
                    missings.append((stud['id'], pair_info['id'], 'н', True))
                elif reason in ('б', 'Б'):
                    missings.append((stud['id'], pair_info['id'], 'б', True))
                elif reason in ('c', 'С', 'с', 'С'):
                    missings.append((stud['id'], pair_info['id'], 'с', False))
                elif reason == '+':
                    pass
            else:
                # var_message(user_id, 'Ничего не меняй в шаблоне, ТОЛЬКО проставь посещаемость')
                return 13
            num += 1
        if missings:
            cur.executemany(
                '''INSERT INTO reports_missingpair(student_id, pair_id, missing_reason, is_debt)
                VALUES (%s, %s, %s, %s)''',
                missings
            )
        cur.execute(
            'UPDATE reports_pairsschedule SET is_reported = %s WHERE id = %s',
            (True, pair_info['id'])
        )
        # connection.commit()
        # var_message(user_id, 'Рапорт принят')
    else:
        # var_message(user_id, 'Ничего не меняй в шаблоне, ТОЛЬКО проставь посещаемость')
        return 14


def test_send_report_excel_pattern(user_id, end_date):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        # var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return 15
    if not (head_info['is_head'] or head_info['is_head_assistant']):
        # var_message(user_id, 'Ты не староста, откуда кнопка😕', kb_params=main_kb_student)
        return 16
    faculty = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    cur.execute(
        '''SELECT min(pr_lk_datetime) as min_date FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s''',
        (faculty, course_n, group_n, False)
    )
    min_date = cur.fetchone()

    if not min_date['min_date']:
        # var_message(user_id, 'Незаполненных рапортов нет', kb_params=main_kb_head)
        return 17

    min_date = min_date['min_date'].replace(tzinfo=None)
    if min_date > end_date:
        # next_p = f'Ближайшая пара {min_date.strftime("%d.%m в %H:%M")}'
        # var_message(user_id, f'На данный момент все рапорты отправлены\n{next_p}', kb_params=main_kb_head)
        return 18

    end_of_min_date_week = (min_date + timedelta(days=6-min_date.weekday())).replace(hour=23, minute=59)
    if end_of_min_date_week < end_date:
        end_date = end_of_min_date_week
    else:
        end_date = end_date
    cur.execute(
        '''SELECT * FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s
        AND pr_lk_datetime BETWEEN %s AND %s
        ORDER BY pr_lk_datetime''',
        (faculty, course_n, group_n, False, min_date, end_date)
    )
    pairs_info = cur.fetchall()
    cur.execute(
        '''SELECT * FROM reports_student
        WHERE faculty_id = %s AND course_n = %s and group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty, course_n, group_n, False)
    )
    group_stud = cur.fetchall()

    wb = Workbook()
    wb_sheet = wb.active
    wb_sheet.cell(row=1, column=1).value = '№'
    wb_sheet.cell(row=1, column=1).border = thin_border
    wb_sheet.merge_cells(start_row=1, end_row=5, start_column=1, end_column=1)
    wb_sheet.column_dimensions['A'].width = 4
    wb_sheet.cell(row=1, column=2).value = 'Фамилия Имя'
    wb_sheet.cell(row=1, column=2).border = thin_border
    wb_sheet.merge_cells(start_row=1, end_row=5, start_column=2, end_column=2)
    stud_names_len = []
    for i, stud in enumerate(group_stud):
        wb_sheet.cell(row=i+6, column=1).value = stud['position_in_group']
        wb_sheet.cell(row=i+6, column=1).border = thin_border
        stud_name = f'{stud["surname"]} {stud["name"]}'
        wb_sheet.cell(row=i+6, column=2).value = stud_name
        wb_sheet.cell(row=i+6, column=2).border = thin_border
        stud_names_len.append(len(stud_name))
    wb_sheet.column_dimensions['B'].width = max(stud_names_len) * 1.2
    start_this_date_col = 3
    extra_cols = 0
    this_date = min_date.date()
    for i, pair_info in enumerate(pairs_info):
        pair_date = pair_info['pr_lk_datetime'].date()
        if this_date != pair_date:
            wb_sheet.cell(row=1, column=start_this_date_col).value = this_date.strftime('%A\n%d.%m.%Y')
            wb_sheet.cell(row=1, column=start_this_date_col).border = thin_border
            wb_sheet.cell(row=1, column=start_this_date_col).alignment = date_alignment
            wb_sheet.merge_cells(start_row=1, start_column=start_this_date_col,
                                 end_row=2, end_column=start_this_date_col + 3 + extra_cols)
            start_this_date_col += 4 + extra_cols
            extra_cols = 0
            this_date = pair_date
        if i + 1 == len(pairs_info):
            wb_sheet.cell(row=1, column=start_this_date_col).value = this_date.strftime('%A\n%d.%m.%Y')
            wb_sheet.cell(row=1, column=start_this_date_col).border = thin_border
            wb_sheet.cell(row=1, column=start_this_date_col).alignment = date_alignment
            wb_sheet.merge_cells(start_row=1, start_column=start_this_date_col,
                                 end_row=2, end_column=start_this_date_col + 3 + extra_cols)
        if pair_info['pair_n'] > 4:
            extra_cols = max(extra_cols, pair_info['pair_n'] - 4)
        this_col = pair_info['pair_n'] + start_this_date_col - 1
        wb_sheet.cell(row=3, column=this_col).value = pair_info['pair_n']
        wb_sheet.cell(row=3, column=this_col).alignment = Alignment(horizontal='center')
        wb_sheet.cell(row=3, column=this_col).border = thin_border

        wb_sheet.cell(row=4, column=this_col).value = 'пр' if pair_info['is_practical'] else 'лк'
        wb_sheet.cell(row=4, column=this_col).alignment = Alignment(horizontal='center')
        wb_sheet.cell(row=4, column=this_col).border = thin_border

        wb_sheet.cell(row=5, column=this_col).value = get_subject_by_pair(pair_info)['short_title']
        wb_sheet.cell(row=5, column=this_col).alignment = Alignment(horizontal='center', textRotation=90)
        wb_sheet.cell(row=5, column=this_col).border = thin_border
        for ii, stud in enumerate(group_stud):
            wb_sheet.cell(row=ii+6, column=this_col).value = choice('++++++++++нбс')
    for col in range(3, wb_sheet.max_column + 1):
        wb_sheet.column_dimensions[wb_sheet.cell(row=3, column=col).column_letter].width = 4
        start_row = 3 if not wb_sheet.cell(row=3, column=col).value else 6
        for r in range(start_row, wb_sheet.max_row + 1):
            wb_sheet.cell(row=r, column=col).alignment = Alignment(horizontal='center')
            wb_sheet.cell(row=r, column=col).border = thin_border

    min_d = min_date.strftime('%d.%m.%Y')
    max_d = end_date.strftime('%d.%m.%Y')
    wb_title = f'{get_faculty_by_id(faculty)["short_title"]} {course_n}курс {group_n}гр рапорт {min_d}-{max_d}.xlsx'
    wb.save(wb_title)
    test_parse_report_excel_pattern(user_id, {'title': wb_title}, end_date)
    # doc_sender(user_id, wb_title, 'Заполни этот рапорт и отправь файл вместе с сообщением "Рапорт"',
    # kb_params=main_kb_head)


def test_parse_report_excel_pattern(user_id, attach, end_date):
    head_info = get_user_by_vk_id(user_id)
    if not head_info:
        # var_message(user_id, 'Ты не авторизован', kb_params=before_login_kb)
        return 19
    if not (head_info['is_head'] or head_info['is_head_assistant']):
        # var_message(user_id, 'Ты не староста и даже не его помощник', kb_params=main_kb_student)
        return 20
    # if not download_attach_xlsx(user_id, attach):
        # os.remove(attach['title'])
        # return 21

    faculty = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    cur.execute(
        '''SELECT min(pr_lk_datetime) as min_date FROM reports_pairsschedule
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s''',
        (faculty, course_n, group_n, False)
    )
    min_date = cur.fetchone()
    if not min_date['min_date']:
        # var_message(user_id, 'Незаполненных рапортов нет, так что этот не приму')
        return 22

    min_date = min_date['min_date'].replace(tzinfo=None)
    if min_date > end_date:
        # next_p = f'Этот рапорт не зачту\nБлижайшая пара {min_date.strftime("%d.%m в %H:%M")}'
        # var_message(user_id, f'На данный момент все рапорты отправлены\n{next_p}')
        return 23

    end_of_min_date_week = (min_date + timedelta(days=6-min_date.weekday())).replace(hour=23, minute=59)
    if end_of_min_date_week < end_date:
        end_date = end_of_min_date_week
    else:
        end_date = end_date
    cur.execute(
        '''SELECT * FROM reports_pairsschedule WHERE
        faculty_id = %s AND course_n = %s AND group_n = %s AND is_reported = %s AND pr_lk_datetime BETWEEN %s AND %s
        ORDER BY pr_lk_datetime''',
        (faculty, course_n, group_n, False, min_date, end_date)
    )
    pairs_info = cur.fetchall()
    cur.execute(
        '''SELECT * FROM reports_student WHERE faculty_id = %s AND course_n = %s and group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty, course_n, group_n, False)
    )
    group_stud = cur.fetchall()

    # [0][0] и [0] пустые чтобы индексы были равны с excel
    wb_matrix = [[], [None, '№', 'Фамилия Имя', None, None, None, None]]
    for _ in range(4):
        wb_matrix.append([None, None, None, None, None, None, None])
    for stud in group_stud:
        wb_matrix.append([None, stud['position_in_group'], f'{stud["surname"]} {stud["name"]}', None, None, None, None])

    start_this_date_col = 3
    extra_cols = 0
    this_date = min_date.date()
    for i, pair_info in enumerate(pairs_info):
        pair_date = pair_info['pr_lk_datetime'].date()
        if this_date != pair_date:
            wb_matrix[1][start_this_date_col] = this_date.strftime('%A\n%d.%m.%Y')
            for index in range(len(wb_matrix)):
                wb_matrix[index] = wb_matrix[index] + [None, None, None, None]
            start_this_date_col += 4 + extra_cols
            extra_cols = 0
            this_date = pair_date
        if i + 1 == len(pairs_info):
            wb_matrix[1][start_this_date_col] = this_date.strftime('%A\n%d.%m.%Y')
        if pair_info['pair_n'] > 4:
            if extra_cols < max(extra_cols, pair_info['pair_n'] - 4):
                new_cols = pair_info['pair_n'] - 4 - extra_cols
                extra_cols = max(extra_cols, pair_info['pair_n'] - 4)
                for index in range(len(wb_matrix)):
                    wb_matrix[index] = wb_matrix[index] + [None for _ in new_cols]
        this_col = pair_info['pair_n'] + start_this_date_col - 1
        wb_matrix[3][this_col] = pair_info['pair_n']
        wb_matrix[4][this_col] = 'пр' if pair_info['is_practical'] else 'лк'
        wb_matrix[5][this_col] = get_subject_by_pair(pair_info)['short_title']

    book = load_workbook(attach['title'])
    sheet = book.active
    is_everything_ok = True
    difference_list = []
    for col in list(sheet.columns)[:2]:
        for cell in col:
            if cell.value != wb_matrix[cell.row][cell.column]:
                is_everything_ok = False
                difference_list.append(f'{wb_matrix[cell.row][cell.column]} на {cell.value}\n')
    for row in list(sheet.rows)[:5]:
        for cell in row[2:]:
            if cell.value != wb_matrix[cell.row][cell.column]:
                is_everything_ok = False
                difference_list.append(f'{wb_matrix[cell.row][cell.column]} на {cell.value}\n')
    if not is_everything_ok:
        difference_list.append('\nЕсли вдруг я оказался неправ, напиши об этом [id39398636|Саше]')
        if len(difference_list) <= 5:
            pass  # var_message(user_id, f'Рапорт не принят\nПохоже ты заменил:\n{"".join(difference_list)}')
        else:
            pass  # var_message(user_id, f'Рапорт не принят\nМного изменений в шаблоне\n{difference_list[-1]}')
        return 24
    attendance_dict = {}
    this_date = min_date.date()
    for col in list(sheet.columns)[2:]:
        if col[0].value:
            this_date = datetime.strptime(col[0].value, '%A\n%d.%m.%Y').date()
            attendance_dict[this_date] = {}
        if col[2].value:
            pair_num = str(col[2].value)
            attendance = {}
            for cell in col[5:]:
                if cell.value:
                    if cell.value not in ('+', 'н', 'Н', 'б', 'Б', 'с', 'С', 'c', 'C'):
                        # var_message(user_id, right_attandance_tags)
                        return 25
                    position = str(sheet.cell(row=cell.row, column=1).value)
                    if cell.value in ('н', 'Н'):
                        attendance[position] = 'н'
                    elif cell.value in ('б', 'Б'):
                        attendance[position] = 'б'
                    elif cell.value in ('c', 'С', 'с', 'C'):
                        attendance[position] = 'с'
                    elif cell.value == '+':
                        attendance[position] = '+'
                else:
                    var_message(user_id, 'Нужно проставить посещаемость всем студентам')
                    return
            attendance_dict[this_date][pair_num] = {
                'pr_or_lk': col[3].value.strip(),
                'subject': col[4].value.strip(),
                'attendance': attendance
            }
    missings = []
    for pair_info in pairs_info:
        pair_date = pair_info['pr_lk_datetime'].date()
        if attendance_dict.get(pair_date):
            pair_n = str(pair_info['pair_n'])
            if attendance_dict[pair_date].get(pair_n):
                cur.execute('UPDATE reports_pairsschedule SET is_reported = %s WHERE id = %s', (True, pair_info['id']))
                pair_n_missings = attendance_dict[pair_date][pair_n]['attendance']
                for stud in group_stud:
                    position = str(stud['position_in_group'])
                    if pair_n_missings.get(position):
                        reason = pair_n_missings[position]
                        if reason != '+':
                            if reason == 'н':
                                missings.append((stud['id'], pair_info['id'], reason, True))
                            elif reason == 'б':
                                missings.append((stud['id'], pair_info['id'], reason, True))
                            elif reason == 'с':
                                missings.append((stud['id'], pair_info['id'], reason, False))
    # print(missings)
    if missings:
        cur.executemany(
            '''INSERT INTO reports_missingpair(student_id, pair_id, missing_reason, is_debt)
            VALUES (%s, %s, %s, %s)''',
            missings
        )
    # var_message(user_id, 'Рапорт принят')
    os.remove(attach['title'])


def fill_reports_random(faculty, course, group_range=None, end_dt=datetime.now()):
    if group_range:
        first_group = int(group_range.split('-')[0])
        last_group = int(group_range.split('-')[1])
        cur.execute(
            '''SELECT * FROM reports_student
            WHERE faculty_id = %s AND course_n = %s AND is_head = %s AND group_n BETWEEN %s AND %s''',
            (faculty, course, True, first_group, last_group)
        )
    else:
        cur.execute(
            'SELECT * FROM reports_student WHERE faculty_id = %s AND course_n = %s AND is_head = %s',
            (faculty, course, True)
        )
    heads = [(dict(x)['group_n'], dict(x)['id'])for x in cur.fetchall()]
    cur.executemany(
        'UPDATE reports_student SET vk_id = %s WHERE id = %s',
        heads
    )
    for head in heads:
        head_id = head[0]
        is_returned = None
        while is_returned is None:
            if choice('++++-') == '+':
                is_returned = test_send_report_message_pattern(head_id, end_dt)
            else:
                is_returned = test_send_report_excel_pattern(head_id, end_dt)
        print(head_id, 'группа -', is_returned)
    cur.executemany(
        'UPDATE reports_student SET vk_id = %s WHERE id = %s',
        [(None, x[1]) for x in heads]
    )
    connection.commit()


# fill_reports_random(1, 4, '19-35', end_dt=datetime(2022, 3, 5))
# check_difference_json_db(1, 4, '19-35')
