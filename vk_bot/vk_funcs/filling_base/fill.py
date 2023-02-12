from collections.abc import Callable
from io import BytesIO

from yadisk import yadisk

from ReportsDjango import settings
from ..config import *
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
import random
import string


# факультеты
def fill_faculty_if_not_exists():
    faculties = [
        ('лечебный', 'Леч'),
        ('педиатрический', 'Пед'),
        ('стоматологический', 'Стом'),
        ('фармацевтический', 'Фарм'),
        ('медико-профилактический', 'МПФ')
    ]
    cur.executemany('INSERT INTO reports_faculty(title, short_title) VALUES (%s, %s) ON CONFLICT DO NOTHING', faculties)


# предметы
def fill_subjects(file_address):
    book = load_workbook(file_address)
    sheet = book.active
    faculty_id = get_faculty_id_by_title(sheet.cell(row=2, column=1).value)
    if not faculty_id:
        if faculty_id is None:
            print('Есть дубль факультетов')
        else:
            print('Такого факультета нет')
        return 0
    course_n = sheet.cell(row=2, column=2).value
    half = sheet.cell(row=2, column=3).value

    row_of_subject = 5
    curations_info = []
    while sheet.cell(row=row_of_subject, column=1).value:
        full_title = sheet.cell(row=row_of_subject, column=1).value
        short_title = sheet.cell(row=row_of_subject, column=2).value
        duration = sheet.cell(row=row_of_subject, column=3).value
        hours_per_pair = sheet.cell(row=row_of_subject, column=4).value
        curations_info.append((full_title, short_title, duration, hours_per_pair, faculty_id, course_n, half))
        row_of_subject += 1

    cur.executemany(
        '''INSERT
        INTO reports_subject(full_title, short_title, duration, hours_per_pair, faculty_id, course_n, half)
        VALUES (%s, %s, %s, %s, %s, %s, %s)''',
        curations_info
    )


# студенты
def fill_students(file_address):
    book = load_workbook(file_address)
    sheet = book.active

    faculty_id = get_faculty_id_by_title(sheet.cell(row=2, column=1).value)
    if not faculty_id:
        if faculty_id is None:
            print('Есть дубль факультетов')
        else:
            print('Такого факультета нет')
        return 0
    course_n = sheet.cell(row=2, column=2).value
    streams = sheet.cell(row=2, column=3).value
    streams = [(int(s_range.split('-')[0]), int(s_range.split('-')[1])) for s_range in streams.split(',')]

    students_list = []
    last_group = 0
    position_in_group = 0
    cur.execute('SELECT * FROM reports_student WHERE id = (SELECT max(id) FROM reports_student)')
    last_student = cur.fetchone()
    if last_student:
        last_student_id = dict(last_student)['id']
    else:
        last_student_id = 0
    start_stud_id = 10000000 + last_student_id
    row_of_student = 5
    while sheet.cell(row=row_of_student, column=2).value:
        surname = sheet.cell(row=row_of_student, column=2).value
        surname = surname.title().strip()
        name = sheet.cell(row=row_of_student, column=3).value
        name = name.title().strip()
        f_name = sheet.cell(row=row_of_student, column=4).value
        if f_name:
            f_name = f_name.title().strip()
        else:
            f_name = ''
        login = 'stud-' + str(start_stud_id+row_of_student-4)[1:]
        password = ''.join((random.choice(string.ascii_letters + '0123456789012345678901234567890123456789') if x != 4 else '-' for x in range(9))).replace('l', 'I')
        group_n = sheet.cell(row=row_of_student, column=1).value
        if last_group == group_n:
            position_in_group += 1
        else:
            last_group = group_n
            position_in_group = 1
        for stream, s_range in enumerate(streams):
            if s_range[0] <= group_n <= s_range[1]:
                stream_n = stream + 1
        is_foreigner = True if sheet.cell(row=row_of_student, column=5).value == '+' else False
        is_head = True if sheet.cell(row=row_of_student, column=6).value == '+' else False
        students_list.append((faculty_id, course_n, stream_n, group_n, surname, name, f_name, is_foreigner, is_head,
                              position_in_group, login, password, not is_head, False, False, '', True, False))
        row_of_student += 1
        students_list.sort(key=lambda x: (x[3], x[9]))
    cur.executemany(
        '''INSERT INTO
        reports_student(faculty_id, course_n, stream_n, group_n, surname, name, f_name, is_foreigner, is_head,
        position_in_group, login, password, is_locked, is_fired, is_head_assistant, can_add_lecture_url,
        get_lec_urls_automatically, is_head_of_head)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
        students_list
    )


# расписание пар и лекций 1-3к
def fill_juniors_schedule(file_address, excel_only=False):
    book = load_workbook(file_address)
    sheet = book.active
    faculty_id = get_faculty_id_by_title(sheet.cell(row=2, column=1).value)
    if not faculty_id:
        print('Такого факультета нет')

    course_n = sheet.cell(row=2, column=2).value
    stream_n = sheet.cell(row=2, column=3).value
    half = sheet.cell(row=2, column=4).value
    start_dt = sheet.cell(row=2, column=5).value
    last_sem_day = sheet.cell(row=2, column=6).value
    last_sem_day = datetime(year=last_sem_day.year, month=last_sem_day.month, day=last_sem_day.day)
    cur.execute(
        '''SELECT * FROM reports_subject WHERE faculty_id = %s AND course_n = %s AND half = %s''',
        (faculty_id, course_n, half)
    )
    all_curations_db = cur.fetchall()
    if all_curations_db:
        all_curations = {dict(x)['short_title']: dict(x) for x in all_curations_db}
    else:
        print('Нет кураций')
        return 0

    # {ДатаВремя: (Факультет, Курс, Поток, ДатаВремя, Предмет, Ссылка)}
    all_lectures = []
    row_to_find_lecture = 5
    while sheet.cell(row=row_to_find_lecture, column=1).value:
        if sheet.cell(row=row_to_find_lecture, column=4).value:
            if sheet.cell(row=row_to_find_lecture, column=4).value.startswith('!ЛЕКЦИЯ!-'):
                for s in sheet.cell(row=row_to_find_lecture, column=4).value[9:].split('///'):
                    lec_time = sheet.cell(row=row_to_find_lecture, column=3).value
                    hours = lec_time.hour
                    minutes = lec_time.minute
                    s = s.split(':')
                    subject = all_curations[s[0]]
                    months = [x.strip() for x in s[1].split(';')]
                    for m in months:
                        year = '20' + m.split('|')[1]
                        month = m.split('|')[0].split('.')[1]
                        dates = m.split('|')[0].split('.')[0].split(',')

                        for d in dates:
                            lec_datetime = datetime.strptime(f'{d}.{month}.{year}', '%d.%m.%Y')\
                                           + timedelta(hours=hours, minutes=minutes)
                            all_lectures.append(
                                (faculty_id, course_n, stream_n, subject['id'], lec_datetime, '', False, None)
                            )
        row_to_find_lecture += 1
    all_lectures.sort(key=lambda x: x[2])
    cur.executemany(
        '''INSERT INTO reports_lecturesschedule(faculty_id, course_n, stream_n, subject_id, lecture_datetime,
        lecture_url, is_sent, stud_add_lec_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
        all_lectures
    )
    cur.execute(
        '''SELECT * FROM reports_lecturesschedule WHERE faculty_id = %s AND course_n = %s AND stream_n = %s''',
        (faculty_id, course_n, stream_n)
    )
    all_lec_db = cur.fetchall()
    all_lectures = {dict(x)['lecture_datetime'].replace(tzinfo=None): dict(x) for x in all_lec_db}

    # {ДеньНедели: {порядок: (время, "!ЛЕКЦИЯ!"), порядок: (время, "!ЛЕКЦИЯ!")}}
    row_of_weekday = 5
    stream_lectures_schedule = {}
    weekday_lectures_schedule = {}
    this_weekday = 0
    while sheet.cell(row=row_of_weekday, column=1).value:
        if this_weekday != sheet.cell(row=row_of_weekday, column=1).value - 1:
            if weekday_lectures_schedule:
                stream_lectures_schedule[this_weekday] = weekday_lectures_schedule
                weekday_lectures_schedule = {}
            this_weekday = sheet.cell(row=row_of_weekday, column=1).value - 1
        this_curation = sheet.cell(row=row_of_weekday, column=4).value
        if this_curation:
            this_curation = this_curation.strip()
            if this_curation.startswith('!ЛЕКЦИЯ!-'):
                pair_num = sheet.cell(row=row_of_weekday, column=2).value
                pair_time = sheet.cell(row=row_of_weekday, column=3).value
                weekday_lectures_schedule[pair_num] = (pair_time, '!ЛЕКЦИЯ!')
        row_of_weekday += 1
    if weekday_lectures_schedule:
        stream_lectures_schedule[this_weekday] = weekday_lectures_schedule

    # Прогон групп
    column_of_group = 4
    stream_n_schedule = {}
    while sheet.cell(row=4, column=column_of_group).value:
        group_n = sheet.cell(row=4, column=column_of_group).value
        row_of_weekday = 5
        this_weekday = 0
        if stream_lectures_schedule.get(this_weekday):
            this_weekday_lectures = stream_lectures_schedule.get(this_weekday)
        else:
            this_weekday_lectures = {}

        # Прогон по строкам группы
        group_n_schedule = {}
        this_weekday_group_n_schedule = {}
        while sheet.cell(row=row_of_weekday, column=1).value:
            # Если следующий день недели
            if this_weekday != sheet.cell(row=row_of_weekday, column=1).value - 1:
                group_n_schedule[this_weekday] = this_weekday_group_n_schedule
                this_weekday_group_n_schedule = {}
                this_weekday = sheet.cell(row=row_of_weekday, column=1).value - 1
                if stream_lectures_schedule.get(this_weekday):
                    this_weekday_lectures = stream_lectures_schedule.get(this_weekday)
                else:
                    this_weekday_lectures = {}
            # Если есть лекция
            pair_num = sheet.cell(row=row_of_weekday, column=2).value
            if this_weekday_lectures.get(pair_num):
                this_lecture_time = this_weekday_lectures.get(pair_num)
                this_weekday_group_n_schedule[pair_num] = (this_lecture_time[0], this_lecture_time[1])
            else:
                this_curation = sheet.cell(row=row_of_weekday, column=column_of_group).value
                if this_curation:
                    this_curation.strip()
                    this_curation = all_curations[this_curation]
                    this_curation_time = sheet.cell(row=row_of_weekday, column=3).value
                    this_weekday_group_n_schedule[pair_num] = (
                        this_curation_time,
                        this_curation['short_title'],
                        this_curation['duration']
                    )
            row_of_weekday += 1
        group_n_schedule[this_weekday] = this_weekday_group_n_schedule
        stream_n_schedule[group_n] = group_n_schedule
        column_of_group += 1

    holidays = get_holidays()

    is_reported = False
    all_groups_schedule = []
    for group_n, week_schedule in stream_n_schedule.items():
        group_n_schedule = []
        subject_counter = {x: 0 for x in all_curations.keys()}
        this_date = start_dt
        while this_date != last_sem_day:
            if this_date.date() not in holidays:
                this_weekday = this_date.date().weekday()
                if week_schedule.get(this_weekday):
                    weekday_schedule = week_schedule[this_weekday]
                    for pair_num in weekday_schedule:
                        this_pair = weekday_schedule[pair_num]
                        if this_pair[1] == '!ЛЕКЦИЯ!':
                            this_lecture_datetime = this_date + timedelta(hours=this_pair[0].hour, minutes=this_pair[0].minute)
                            is_practical = False
                            if all_lectures.get(this_lecture_datetime):
                                this_lecture = all_lectures[this_lecture_datetime]
                                group_n_schedule.append(
                                    (faculty_id, course_n, group_n, is_practical, this_lecture['subject_id'],
                                     this_lecture['id'], this_lecture_datetime, pair_num, is_reported, '')
                                )
                        else:
                            is_practical = True
                            pair_subject = this_pair[1]
                            this_pair_datetime = this_date + timedelta(hours=this_pair[0].hour, minutes=this_pair[0].minute)
                            if subject_counter[pair_subject] != this_pair[2]:
                                subject_counter[pair_subject] = subject_counter[pair_subject] + 1
                                group_n_schedule.append(
                                    (faculty_id, course_n, group_n, is_practical, all_curations[pair_subject]['id'],
                                     None, this_pair_datetime, pair_num, is_reported, '')
                                )
            this_date += timedelta(days=1)
        all_groups_schedule.extend(group_n_schedule)

    if not excel_only:
        cur.executemany(
            '''INSERT INTO reports_pairsschedule(faculty_id, course_n, group_n, is_practical, subject_id,
            lecture_id, pr_lk_datetime, pair_n, is_reported, pr_lk_comment)
            VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
            all_groups_schedule
        )
        # connection.commit()
    else:
        all_curations_by_id = {dict(x)['id']: dict(x)['short_title'] for x in all_curations_db}
        all_lec_by_id = {dict(x)['id']: dict(x) for x in all_lec_db}

        wb = Workbook()
        wb_sheet = wb.active
        wb_sheet.cell(row=1, column=1).value = 'факультет'
        wb_sheet.cell(row=1, column=2).value = 'курс'
        wb_sheet.cell(row=1, column=3).value = 'группа'
        wb_sheet.cell(row=1, column=4).value = 'пр?'
        wb_sheet.cell(row=1, column=5).value = 'предмет'
        wb_sheet.cell(row=1, column=6).value = 'лекция'
        wb_sheet.cell(row=1, column=7).value = 'ДатаВремя'
        wb_sheet.cell(row=1, column=8).value = '№ пары'
        wb_sheet.cell(row=1, column=9).value = 'есть рапорт'
        wb_sheet.cell(row=1, column=10).value = 'ссылка'
        for i, note in enumerate(all_groups_schedule):
            wb_sheet.cell(row=i+2, column=1).value = sheet.cell(row=2, column=1).value
            wb_sheet.cell(row=i+2, column=2).value = note[1]
            wb_sheet.cell(row=i+2, column=3).value = note[2]
            wb_sheet.cell(row=i+2, column=4).value = note[3]
            wb_sheet.cell(row=i+2, column=5).value = all_curations_by_id[note[4]]
            wb_sheet.cell(row=i+2, column=6).value = all_curations_by_id[all_lec_by_id[note[5]]['subject_id']] if note[5] else None
            wb_sheet.cell(row=i+2, column=7).value = note[6]
            wb_sheet.cell(row=i+2, column=8).value = note[7]
            wb_sheet.cell(row=i+2, column=9).value = note[8]
            wb_sheet.cell(row=i+2, column=10).value = note[9]

        new_file_address = '/'.join(['ГОТОВОЕ ' + x for x in file_address.split('/') if '.xlsx' in x])
        wb.save(new_file_address)


# Расписание пар и лекций 4-6к
def fill_elders_schedule(file_address, excel_only=False):
    book = load_workbook(file_address)
    sheet = book.active
    faculty_id = get_faculty_id_by_title(sheet.cell(row=2, column=1).value)
    if not faculty_id:
        if faculty_id is None:
            print('Есть дубль факультетов')
        else:
            print('Такого факультета нет')
        return 0
    course_n = sheet.cell(row=2, column=2).value
    stream_n = sheet.cell(row=2, column=3).value
    half = sheet.cell(row=2, column=4).value
    fks_weekday = sheet.cell(row=2, column=5).value
    fks_time = sheet.cell(row=2, column=6).value
    curs_time = sheet.cell(row=2, column=7).value
    start_dt = sheet.cell(row=2, column=8).value

    cur.execute(
        'SELECT * FROM reports_subject WHERE faculty_id = %s AND course_n = %s AND half = %s',
        (faculty_id, course_n, half)
    )
    all_curations_db = cur.fetchall()
    all_curations = {}
    if all_curations_db:
        for curation in all_curations_db:
            all_curations[curation['short_title']] = curation
    else:
        print('Загрузи пары для этого курса')

    lec_schedule = []
    month_col = 4
    while sheet.cell(row=4, column=month_col).value:
        row_of_month = 5
        lec_month = sheet.cell(row=4, column=month_col).value
        while sheet.cell(row=row_of_month, column=2).value:
            lec_time = sheet.cell(row=row_of_month, column=2).value
            lec_subject = sheet.cell(row=row_of_month, column=3).value
            if all_curations.get(lec_subject):
                subject = all_curations.get(lec_subject)
            else:
                print(f'Нет предмета {lec_subject} в базе данных\nЛекции, строка {row_of_month}')
                return 0
            if sheet.cell(row=row_of_month, column=month_col).value:
                for day in str(sheet.cell(row=row_of_month, column=month_col).value).split('_'):
                    lec_datetime = lec_month + timedelta(days=int(day)-1, hours=lec_time.hour, minutes=lec_time.minute)
                    lec_schedule.append(
                        (faculty_id, course_n, stream_n, subject['id'], lec_datetime, '', False, None)
                    )
            row_of_month += 1
        month_col += 1

    lec_schedule.sort(key=lambda x: x[4])
    cur.executemany(
        '''INSERT INTO reports_lecturesschedule(faculty_id, course_n, stream_n, subject_id, lecture_datetime,
        lecture_url, is_sent, stud_add_lec_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
        lec_schedule
    )
    cur.execute(
        '''SELECT * FROM reports_lecturesschedule WHERE faculty_id = %s AND course_n = %s AND stream_n = %s''',
        (faculty_id, course_n, stream_n)
    )
    all_lec_db = cur.fetchall()
    all_lectures = {dict(x)['lecture_datetime'].replace(tzinfo=None): dict(x) for x in all_lec_db}

    holidays = get_holidays()
    weekends = (5, 6)

    stream_n_schedule = []
    group_col = 13
    fks_info = all_curations['ФКС']
    while sheet.cell(row=4, column=group_col).value:
        group_n = sheet.cell(row=4, column=group_col).value
        this_cur_row = 5
        this_date = start_dt
        fks_count = 0
        group_n_schedule = []
        while sheet.cell(row=this_cur_row, column=group_col).value:
            cur_subject = sheet.cell(row=this_cur_row, column=group_col).value
            if all_curations.get(cur_subject):
                cur_subject = all_curations[cur_subject]
            else:
                print(f'Нет предмета {cur_subject} в базе данных\nКурация, группа {group_n}')
                return 0
            for cur_num in range(cur_subject['duration']):
                this_date_group_schedule = []
                while this_date.date() in holidays or this_date.weekday() in weekends:
                    this_date = this_date + timedelta(days=1)
                # ФКС
                if this_date.weekday() == fks_weekday:
                    if fks_count < fks_info['duration']:
                        fks_datetime = this_date + timedelta(hours=fks_time.hour, minutes=fks_time.minute)
                        this_date_group_schedule.append([
                            faculty_id, course_n, group_n, True, fks_info['id'], None, fks_datetime, False, ''
                        ])
                        fks_count += 1
                # Лекции
                for lec_datetime in all_lectures:
                    if lec_datetime.date() == this_date.date():
                        this_lec = all_lectures[lec_datetime]
                        this_date_group_schedule.append([
                            faculty_id, course_n, group_n, False, this_lec['subject_id'], this_lec['id'], lec_datetime, False, ''
                        ])
                # Пара
                cur_datetime = this_date + timedelta(hours=curs_time.hour, minutes=curs_time.minute)
                this_date_group_schedule.append([
                    faculty_id, course_n, group_n, True, cur_subject['id'], None, cur_datetime, False, ''
                ])
                this_date_group_schedule.sort(key=lambda x: x[6])
                for pair_n, pair in enumerate(this_date_group_schedule):
                    pair.insert(7, pair_n + 1)
                    group_n_schedule.append(tuple(pair))
                this_date = this_date + timedelta(days=1)
            this_cur_row += 1
        if this_date.date() <= list(all_lectures.keys())[-1].date():
            lectures_over_curations = {}
            for lec_datetime in all_lectures:
                if lec_datetime.date() > this_date.date():
                    this_lec = all_lectures[lec_datetime]
                    if lectures_over_curations.get(lec_datetime.date()):
                        x = lectures_over_curations[lec_datetime.date()]
                        x.append((
                            faculty_id, course_n, group_n, False, this_lec['subject_id'], this_lec['id'], lec_datetime,
                            [1 for _ in x].count(1) + 1, False, ''
                        ))
                        lectures_over_curations[lec_datetime.date()] = x
                    else:
                        lectures_over_curations[lec_datetime.date()] = [(
                            faculty_id, course_n, group_n, False, this_lec['subject_id'], this_lec['id'],
                            lec_datetime, 1, False, '')]
            for lec in lectures_over_curations.values():
                group_n_schedule.extend(lec)

        stream_n_schedule.extend(group_n_schedule)
        group_col += 1

    if not excel_only:
        cur.executemany(
            '''INSERT INTO reports_pairsschedule(faculty_id, course_n, group_n, is_practical, subject_id,
            lecture_id, pr_lk_datetime, pair_n, is_reported, pr_lk_comment)
            VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
            stream_n_schedule
        )
        # connection.commit()
    else:
        all_curations_by_id = {dict(x)['id']: dict(x)['short_title'] for x in all_curations_db}
        all_lec_by_id = {dict(x)['id']: dict(x) for x in all_lec_db}

        wb = Workbook()
        wb_sheet = wb.active
        wb_sheet.cell(row=1, column=1).value = 'факультет'
        wb_sheet.cell(row=1, column=2).value = 'курс'
        wb_sheet.cell(row=1, column=3).value = 'группа'
        wb_sheet.cell(row=1, column=4).value = 'пр?'
        wb_sheet.cell(row=1, column=5).value = 'предмет'
        wb_sheet.cell(row=1, column=6).value = 'лекция'
        wb_sheet.cell(row=1, column=7).value = 'ДатаВремя'
        wb_sheet.cell(row=1, column=8).value = '№ пары'
        wb_sheet.cell(row=1, column=9).value = 'есть рапорт'
        wb_sheet.cell(row=1, column=10).value = 'ссылка'
        for i, note in enumerate(stream_n_schedule):
            wb_sheet.cell(row=i+2, column=1).value = sheet.cell(row=2, column=1).value
            wb_sheet.cell(row=i+2, column=2).value = note[1]
            wb_sheet.cell(row=i+2, column=3).value = note[2]
            wb_sheet.cell(row=i+2, column=4).value = note[3]
            wb_sheet.cell(row=i+2, column=5).value = all_curations_by_id[note[4]]
            wb_sheet.cell(row=i+2, column=6).value = all_curations_by_id[all_lec_by_id[note[5]]['subject_id']] if note[5] else None
            wb_sheet.cell(row=i+2, column=7).value = note[6]
            wb_sheet.cell(row=i+2, column=8).value = note[7]
            wb_sheet.cell(row=i+2, column=9).value = note[8]
            wb_sheet.cell(row=i+2, column=10).value = note[9]

        new_file_address = '/'.join(['ГОТОВОЕ ' + x for x in file_address.split('/') if '.xlsx' in x])
        wb.save(new_file_address)


def fill_by_files_in_dirs(subjects=False, students=False, schedules_e=False, schedules_j=False, **kwargs):
    ya = yadisk.YaDisk(token=settings.YADISK_TOKEN)

    def fill(folder: str, handler: Callable) -> None:
        for file in ya.listdir('/СУП excels/' + folder):
            resp = requests.get(file.file)
            resp.raise_for_status()
            handler(BytesIO(resp.content))
            print(file.name + ' ----> успешно загружен')

    fill_faculty_if_not_exists()
    if subjects:
        fill('subjects', fill_subjects)
    if students:
        fill('students', fill_students)
    if schedules_j:
        fill('schedules/1,2,3', fill_juniors_schedule)
    if schedules_e:
        fill('schedules/4,5,6', fill_elders_schedule)


def truncate_db():
    cur.execute(
        'TRUNCATE '
        'reports_changesattendance, '
        'reports_lecturesschedule, '
        'reports_missingpair, '
        'reports_pairsschedule, '
        'reports_student, '
        'reports_studentsdoc, '
        'reports_studentsvaccine, '
        'reports_subject '
        'RESTART IDENTITY'
    )


def get_holidays():
    return [
        datetime.strptime('23.02.2023', '%d.%m.%Y').date(),
        datetime.strptime('8.03.2023', '%d.%m.%Y').date(),
        datetime.strptime('1.05.2023', '%d.%m.%Y').date(),
        datetime.strptime('9.05.2023', '%d.%m.%Y').date(),
    ]
