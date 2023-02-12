from pathlib import Path

from openpyxl import Workbook

from ..config import *
from ..excels_styles import *


def generate(course: int, folder_to_save: Path):
    if not folder_to_save.exists():
        folder_to_save.mkdir()

    cur.execute(
        '''SELECT group_n, surname, name, login, password FROM reports_student
        WHERE faculty_id = %s AND course_n = %s AND is_head = %s''',
        (1, course, True)
    )
    heads = cur.fetchall()
    if not heads:
        print(f"{course} курс - нет старост")
        return

    wb = Workbook()
    sheet = wb.active
    sheet.cell(1, 1).value = 'Группа'
    sheet.cell(1, 2).value = 'Фамилия'
    sheet.cell(1, 3).value = 'Имя'
    sheet.cell(1, 4).value = 'Логин'
    sheet.cell(1, 5).value = 'Пароль'
    surnames_len = []
    names_len = []
    logins_len = []
    passwords_len = []
    for i, head in enumerate(heads):
        this_row = i + 2
        sheet.cell(this_row, 1).value = head['group_n']
        sheet.cell(this_row, 2).value = head['surname']
        sheet.cell(this_row, 3).value = head['name']
        sheet.cell(this_row, 4).value = head['login']
        sheet.cell(this_row, 5).value = head['password']
        surnames_len.append(len(head['surname']))
        names_len.append(len(head['name']))
        logins_len.append(len(head['login']))
        passwords_len.append(len(head['password']))
    sheet.column_dimensions['A'].width = 7
    for c in (('B', max(surnames_len)), ('C', max(names_len)), ('D', max(logins_len)), ('E', max(passwords_len))):
        sheet.column_dimensions[c[0]].width = c[1] * 1.2
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column + 1):
            sheet.cell(i, j).border = top_double_border
    passwords_file = folder_to_save / f'Пароли старост {course} курс.xlsx'
    wb.save(passwords_file)
    doc_sender(VK_GROUP_OWNER_ID, passwords_file)
    passwords_file.unlink()  # удаляем
