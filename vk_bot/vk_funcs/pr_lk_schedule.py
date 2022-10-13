from .config import *
from .all_buttons import *
from .excels_styles import *
from datetime import datetime, timedelta
from openpyxl import Workbook


def lecture_url_add_or_change_payload_processing(user_id, stud_info, payload_par: str):
    payload_par = payload_par.split('-')
    add = True if payload_par[0] == 'add' else False
    if len(payload_par) == 1:
        streams = [int(x) for x in stud_info['can_add_lecture_url'].split(',')]
        if len(streams) == 1:
            show_subjects_to_add_or_change_lecture_url(user_id, stud_info, streams[0], add)
        else:
            num = 1
            streams_kb = {}
            mode = 'add' if add else 'change'
            for s in streams:
                streams_kb[num] = [(f'{s} поток', 'primary', f'{{"button":"schedule_lecture_url_{mode}-{s}"}}'), ]
                num += 1
            streams_kb[num] = (cancel_btn,)
            var_message(user_id, '...', kb_params=streams_kb)
    elif len(payload_par) == 2:
        show_subjects_to_add_or_change_lecture_url(user_id, stud_info, int(payload_par[1]), add)
    elif len(payload_par) == 3:
        show_dates_to_add_or_change_lecture_url(user_id, stud_info, int(payload_par[1]), int(payload_par[2]), add)
    elif len(payload_par) == 4:
        send_lecture_url_add_or_change_pattern(user_id, stud_info, int(payload_par[1]), int(payload_par[3]), add)


def show_subjects_to_add_or_change_lecture_url(user_id, stud_info, lec_stream, add=True):
    faculty_id = stud_info['faculty_id']
    course_n = stud_info['course_n']

    cur.execute(
        f'''SELECT
            ls.subject_id AS subject_id,
            s.short_title AS subject
        FROM reports_lecturesschedule AS ls
            JOIN reports_subject AS s
            ON s.id = ls.subject_id
        WHERE
            ls.faculty_id = %s AND
            ls.course_n = %s AND
            ls.stream_n = %s AND
            ls.lecture_url {"" if add else "!"}= %s AND
            ls.lecture_datetime > %s
        GROUP BY s.short_title, ls.subject_id''',
        (faculty_id, course_n, lec_stream, '', datetime.now().replace(hour=0, minute=0, second=0))
    )
    subjects = cur.fetchall()
    if not subjects:
        var_message(user_id, f'Лекций {"без ссылок больше" if add else "с ссылками пока"} нет',
                    main_kb_head if stud_info['is_head'] or stud_info['is_head_assistant'] else main_kb_student)
        return

    num = 1
    subj_schedule_kb, subj_btn_line = {num: []}, []
    mode = "add" if add else "change"
    payload_button = f'{{{{"button": "schedule_lecture_url_{mode}-{lec_stream}-{{}}"}}}}'
    msx_btn_in_line = 3 if len(subjects) > 16 else 2
    for s in subjects:
        subj_schedule_kb[num].append((s['subject'], 'primary', payload_button.format(s['subject_id'])))
        if len(subj_schedule_kb[num]) == msx_btn_in_line:
            num += 1
            subj_schedule_kb[num] = []
    if subj_schedule_kb[num]:
        num += 1
    subj_schedule_kb[num] = (('<-Назад', 'negative', f'{{"button":"schedule_lecture_url_{mode}"}}'), cancel_btn)
    var_message(user_id, '...', kb_params=subj_schedule_kb)


def show_dates_to_add_or_change_lecture_url(user_id, stud_info, lec_stream, lec_subject_id, add=True):
    faculty_id = stud_info['faculty_id']
    course_n = stud_info['course_n']

    cur.execute(
        f'''SELECT
            id AS id,
            lecture_datetime AS lec_dt
        FROM reports_lecturesschedule
        WHERE
            faculty_id = %s AND
            course_n = %s AND
            stream_n = %s AND
            subject_id = %s AND
            lecture_datetime > %s AND
            lecture_url {"" if add else "!"}= %s
        ORDER BY lecture_datetime
        FETCH FIRST 19 ROW ONLY''',
        (faculty_id, course_n, lec_stream, lec_subject_id, datetime.now().replace(hour=0, minute=0, second=0), '')
    )
    lectures = cur.fetchall()
    if not lectures:
        var_message(user_id, f'Лекций {"без ссылок больше" if add else "с ссылками пока"} нет',
                    main_kb_head if stud_info['is_head'] or stud_info['is_head_assistant'] else main_kb_student)
        return

    num = 1
    lecture_dates_kb, date_btn_line = {num: []}, []
    mode = "add" if add else "change"
    payload_button = f'{{{{"button": "schedule_lecture_url_{mode}-{lec_stream}-{lec_subject_id}-{{}}"}}}}'
    max_btn_in_line = 3 if len(lectures) > 16 else 2
    for lec in lectures:
        lecture_dates_kb[num].append(
            (lec['lec_dt'].strftime('%d.%m %H:%M'), 'primary', payload_button.format(lec['id']))
        )
        if len(lecture_dates_kb[num]) == max_btn_in_line:
            num += 1
            lecture_dates_kb[num] = []
    if lecture_dates_kb[num]:
        num += 1
    lecture_dates_kb[num] = (('<-Назад', 'negative', f'{{"button":"schedule_lecture_url_{mode}-{lec_stream}"}}'), cancel_btn)
    var_message(user_id, '...', lecture_dates_kb)


def send_lecture_url_add_or_change_pattern(user_id, stud_info, lec_stream, lec_id, add=True, url=None):
    faculty_id = stud_info['faculty_id']
    course_n = stud_info['course_n']

    cur.execute(
        f'''SELECT
            ls.stud_add_lec_id AS stud_id,
            s.short_title AS subject,
            ls.lecture_datetime AS lec_dt,              
            ls.lecture_url AS url
        FROM reports_lecturesschedule AS ls
            JOIN reports_subject AS s
            on s.id = ls.subject_id
        WHERE
            ls.faculty_id = %s AND
            ls.course_n = %s AND
            ls.stream_n = %s AND
            ls.id = %s''',
        (faculty_id, course_n, lec_stream, lec_id)
    )
    this_lec = cur.fetchone()

    if not this_lec:
        var_message(user_id, 'Такой лекции нет, такого быть не должно, напиши об этом [id39398636|Саше]')
        return

    lec_dt = this_lec['lec_dt'].replace(tzinfo=None).strftime('%d.%m.%Y в %H:%M')
    if add:
        if this_lec['url'] != '':
            var_message(user_id,
                        f'Ссылка на лекцию для {lec_stream} потока по {this_lec["subject"]} {lec_dt} уже заполнена')
            return
    else:
        if this_lec['url'] == '':
            var_message(user_id,
                        f'Ссылка на лекцию для {lec_stream} потока по {this_lec["subject"]} {lec_dt} ещё не заполнена')
            return

    mode = 'ДОБАВИТЬ' if add else 'ИЗМЕНИТЬ'
    pattern = f'Ссылка {mode}\nПоток - {lec_stream}\nПредмет - {this_lec["subject"]}\nВремя - {lec_dt}\n' \
              f'{{{url if url else ""}}}'
    var_message(user_id, pattern)

    if not add:
        cur.execute('SELECT * FROM reports_student WHERE id = %s', (this_lec['stud_id'], ))
        stud_add_url = cur.fetchone()
        last_url = ''
        if not url:
            if stud_add_url:
                last_url = f'\n\nТекущая ссылка (добавил {stud_add_url["name"]} {stud_add_url["surname"]})\n' \
                           f'{this_lec["url"]}'
    else:
        last_url = ''
    if not url:
        var_message(user_id, 'Добавь новую ссылку внутрь фигурных скобок или отправь ссылку с вложенным сообщением '
                             'с шаблоном' + last_url)


def parse_lecture_url_add_or_change_message(user_id, stud_info, msg, add, recursive):
    faculty_id = stud_info['faculty_id']
    course_n = stud_info['course_n']

    if stud_info['can_add_lecture_url']:
        streams = [int(x) for x in stud_info['can_add_lecture_url'].split(',')]
    else:
        var_message(user_id, 'Ты не можешь добавлять ссылки на лекции')
        return

    msg = msg.split('\n')

    if len(msg) >= 5:
        if msg[1].startswith('Поток -'):
            try:
                stream_n = int(msg[1].split('-')[1])
                if stream_n not in streams:
                    var_message(user_id, 'Ты не можешь добавлять ссылки на лекции для этого потока')
                    return
            except IndexError:
                var_message(user_id, 'Не заполнен поток')
                return
            except ValueError:
                var_message(user_id, 'По-моему в потоке не цифра\nНе изменяй шаблон')
                return
        else:
            var_message(user_id, 'Не изменяй шаблон (поток)')
            return

        if msg[2].startswith('Предмет -'):
            try:
                subject = get_subject_by_title(faculty_id, course_n, half=1, subject_title=msg[2].split('-')[1].strip())
                if not subject:
                    var_message(user_id, 'Такого предмета нет')
                    return
            except IndexError:
                var_message(user_id, 'Предмет не заполнен')
                return
        else:
            var_message(user_id, 'Не изменяй шаблон (предмет)')
            return

        if msg[3].startswith('Время -'):
            try:
                lec_dt = datetime.strptime(msg[3].split('-')[1].strip(), '%d.%m.%Y в %H:%M')
                if lec_dt < datetime.now().replace(hour=0, minute=0):
                    var_message(user_id, 'Уже поздно добавлять ссылку')
                    return
            except IndexError:
                var_message(user_id, 'Время не заполнено')
                return
            except ValueError:
                var_message(user_id, 'Время написано неправильно')
                return
        else:
            var_message(user_id, 'Не изменяй шаблон (время)')
            return

        if msg[4].startswith('{') and msg[4].endswith('}'):
            lec_url = msg[4][1:-1].strip()
            if lec_url == '':
                if add:
                    var_message(user_id, 'Ссылка не заполнена')
                    return
                else:
                    lec_url = ''
            elif len(lec_url) > 490:
                var_message(user_id, 'Ссылка слишком длинная')
                return
        else:
            lec_url = '\n'.join(msg[4:])
            if lec_url.startswith('{') and lec_url.endswith('}'):
                lec_url = lec_url[1:-1].strip()
            else:
                var_message(user_id, 'Ссылка должна быть в фигурных скобках')
                return
    else:
        var_message(user_id, 'Мало строк, не меняй шаблон')
        return

    sql_req_pars = [faculty_id, course_n, stream_n, subject['id'], lec_dt]
    if add:
        sql_req_pars = [faculty_id, course_n, stream_n, subject['id'], lec_dt, '']
    cur.execute(
        f'''SELECT * 
        FROM reports_lecturesschedule
        WHERE
            faculty_id = %s AND
            course_n = %s AND
            stream_n = %s AND
            subject_id = %s AND
            lecture_datetime = %s {" AND lecture_url = %s" if add else ""}''',
        sql_req_pars
    )
    this_lec = cur.fetchone()
    if not this_lec:
        var_message(user_id, 'Такой лекции нет или ссылка уже добавлена')
        return
    if add:
        if this_lec['lecture_url'] != '':
            var_message(user_id, f'На эту лекцию уже была добавлена ссылка\n\n{this_lec["lecture_url"]}')
            return
        mode = 'добавлена'
    else:
        mode = 'изменена'
    cur.execute(
        '''UPDATE reports_lecturesschedule SET lecture_url = %s, stud_add_lec_id = %s WHERE id = %s''',
        (lec_url, stud_info['id'], this_lec['id'])
    )

    if msg[4].endswith('}'):
        var_message(user_id,
                    f'Ссылка на лекцию для {stream_n} потока по {subject["short_title"]} {lec_dt} {mode}\n\n{lec_url}')
    else:
        var_message(user_id, f'Ссылка на лекцию для {stream_n} потока по {subject["short_title"]} {lec_dt} {mode}\n\n'
                             f'{lec_url}\n\n'
                             'В ссылке были переносы строк, я добавил как есть\nЕсли нужно поменять ссылку на такую же,'
                             ' только без переносов, просто скопируй и отправь следующее сообщение')
        new_lec_url = lec_url.replace('\n', '')
        send_lecture_url_add_or_change_pattern(user_id, stud_info, stream_n, this_lec['id'], False, new_lec_url)

    if this_lec['lecture_datetime'].replace(tzinfo=None) < datetime.now() + timedelta(minutes=30):
        send_lectures_automatically(this_lec['id'])


def show_lectures_for_two_more_days(user_id, stud_info):
    faculty_id = stud_info['faculty_id']
    course_n = stud_info['course_n']
    stream_n = stud_info['stream_n']

    begin_time = datetime.now().replace(hour=0, minute=0)
    end_time = begin_time + timedelta(days=2)

    def get_stream_lectures(str_n):
        cur.execute(
            '''SELECT
                s.short_title AS subject,
                ls.lecture_datetime AS lec_dt,
                ls.lecture_url AS lec_url
            FROM reports_lecturesschedule AS ls
                JOIN reports_subject AS s
                ON s.id = ls.subject_id
            WHERE
                ls.faculty_id = %s AND
                ls.course_n = %s AND
                ls.stream_n = %s AND
                ls.lecture_datetime BETWEEN %s AND %s
            ORDER BY
                lec_dt''',
            (faculty_id, course_n, str_n, begin_time, end_time)
        )
        two_days_lecs = cur.fetchall()

        if not two_days_lecs:
            return 'Сегодня и завтра нет лекций'

        last_day = two_days_lecs[0]['lec_dt'].date()
        lectures_info = f'{last_day.strftime("%d.%m.%Y")}\n'
        for lec in two_days_lecs:
            this_day = lec['lec_dt'].date()
            if this_day != last_day:
                lectures_info += f'\n{this_day.strftime("%d.%m.%Y")}\n'
            lec_url = lec["lec_url"] if lec["lec_url"] != '' else '(Ссылки нет)'
            lectures_info += f'{lec["subject"]} - {lec["lec_dt"].strftime("%H.%M")}\n{lec_url}\n'
            last_day = this_day
        return lectures_info

    if stud_info['can_add_lecture_url']:
        streams = [int(x) for x in stud_info['can_add_lecture_url'].split(',')]
        if not stream_n in streams:
            streams.append(stream_n)
            streams.sort()
        reply_msgs = []
        for stream in streams:
            reply_msgs.append(f'{stream} поток:\n{get_stream_lectures(stream)}')
        for reply_msg in reply_msgs:
            var_message(user_id, reply_msg)
    else:
        reply_msg = get_stream_lectures(stream_n)
        var_message(user_id, reply_msg,
                    main_kb_head if stud_info['is_head'] or stud_info['is_head_assistant'] else main_kb_student)


def show_week_schedule(user_id, stud_info, payload_pars: str):
    payload_pars = payload_pars.split('-')
    if len(payload_pars) == 1:
        var_message(user_id, '...', kb_params=show_schedule_kb)
    elif len(payload_pars) == 2:
        faculty_id = stud_info['faculty_id']
        course_n = stud_info['course_n']
        group_n = stud_info['group_n']

        this_week_begin_dt = datetime.now().replace(hour=0, minute=0) - timedelta(days=datetime.now().weekday())
        if payload_pars[1] == 'this':
            week_begin_dt = this_week_begin_dt
        elif payload_pars[1] == 'next':
            week_begin_dt = this_week_begin_dt + timedelta(days=7)
        else:  # if payload_pars[1] == 'last'
            week_begin_dt = this_week_begin_dt - timedelta(days=7)
        week_end_dt = (week_begin_dt + timedelta(days=6)).replace(hour=23, minute=59)

        cur.execute(
            '''SELECT
                ps.pr_lk_datetime AS pair_dt,
                ps.pair_n AS pair_n,
                s.short_title AS subject,
                ps.is_practical AS is_pr,
                ps.pr_lk_comment AS pr_com,
                ls.lecture_url AS lec_url
            FROM reports_pairsschedule AS ps
                JOIN reports_subject AS s
                ON s.id = ps.subject_id
                LEFT JOIN reports_lecturesschedule AS ls
                ON ls.id = ps.lecture_id
            WHERE
                ps.faculty_id = %s AND
                ps.course_n = %s AND
                ps.group_n = %s AND
                ps.pr_lk_datetime BETWEEN %s AND %s
            ORDER BY ps.pr_lk_datetime''',
            (faculty_id, course_n, group_n, week_begin_dt, week_end_dt)
        )
        week_schedule = cur.fetchall()
        if not week_schedule:
            var_message(user_id, f'Нет пар с {week_begin_dt.strftime("%d.%m")} по {week_end_dt.strftime("%d.%m")}')
            return

        week_schedule_dict = {}
        for pair in week_schedule:
            pair_date = pair['pair_dt'].date()
            if week_schedule_dict.get(pair_date) is None:
                week_schedule_dict[pair_date] = []
            week_schedule_dict[pair_date].append(
                (pair['pair_n'], pair['is_pr'], pair['subject'], pair['pr_com'], pair['lec_url'])
            )
        week_schedule_sorted_list = [(d, p) for d, p in week_schedule_dict.items()]
        week_schedule_sorted_list.sort(key=lambda x: x[0])

        wb = Workbook()
        wb_sheet = wb.active
        this_row = 0
        ex_col = 0
        wb_sheet.column_dimensions['A'].width = 15
        wb_sheet.column_dimensions['B'].width = 3
        wb_sheet.column_dimensions['C'].width = 4
        wb_sheet.column_dimensions['E'].width = 100
        for pair_date, pairs in week_schedule_sorted_list:
            wb_sheet.cell(row=this_row+1, column=1).value = pair_date.strftime('%A\n%d.%m.%Y')
            wb_sheet.cell(row=this_row+1, column=1).alignment = date_alignment
            if max([p[0] for p in pairs]) > 4:
                ex_col = max([p[0] for p in pairs]) - 4
            wb_sheet.merge_cells(start_row=this_row+1, end_row=this_row + 4 + ex_col, start_column=1, end_column=1)
            for pair in pairs:
                wb_sheet.cell(row=this_row+pair[0], column=2).value = pair[0]
                wb_sheet.cell(row=this_row+pair[0], column=4).value = pair[2]
                if pair[1]:
                    wb_sheet.cell(row=this_row+pair[0], column=3).value = 'пр'
                    wb_sheet.cell(row=this_row+pair[0], column=5).value = pair[3]
                else:
                    wb_sheet.cell(row=this_row+pair[0], column=3).value = 'лк'
                    wb_sheet.cell(row=this_row+pair[0], column=5).value = pair[4] if pair[4] != '' else 'Ссылки ещё нет'
            this_row += 4 + ex_col
            ex_col = 0
        for i in range(1, wb_sheet.max_row + 1):
            for j in range(1, wb_sheet.max_column + 1):
                wb_sheet.cell(row=i, column=j).border = thin_border
        wb_title = f'Расписание {get_faculty_by_id(faculty_id)["short_title"]} {course_n} курс {group_n} группа'
        wb_title += f' {week_begin_dt.strftime("%d.%m.%Y")} - {week_end_dt.strftime("%d.%m.%Y")}.xlsx'
        wb.save(wb_title)
        doc_sender(user_id, wb_title,
                   kb_params=main_kb_head if stud_info['is_head'] or stud_info['is_head_assistant'] else main_kb_student)


def send_lectures_automatically(lec_id=None):
    now_dt = datetime.now()
    next_30_min = now_dt + timedelta(minutes=30)
    if lec_id:
        cur.execute(
            '''SELECT
                ls.id AS id,
                ls.faculty_id AS faculty_id,
                ls.course_n AS course_n,
                ls.stream_n AS stream_n,
                s.short_title AS subject,
                ls.lecture_datetime AS lec_dt,
                ls.lecture_url AS lec_url
            FROM reports_lecturesschedule AS ls
                JOIN reports_subject AS s
                ON s.id = ls.subject_id
            WHERE
                ls.id = %s''',
            (lec_id, )
        )
    else:
        cur.execute(
            '''SELECT
                ls.id AS id,
                ls.faculty_id AS faculty_id,
                ls.course_n AS course_n,
                ls.stream_n AS stream_n,
                s.short_title AS subject,
                ls.lecture_datetime AS lec_dt,
                ls.lecture_url AS lec_url
            FROM reports_lecturesschedule AS ls
                JOIN reports_subject AS s
                ON s.id = ls.subject_id
            WHERE
                lecture_datetime BETWEEN %s AND %s AND
                is_sent = %s''',
            (now_dt, next_30_min, False)
        )
    lectures = cur.fetchall()
    if lectures:
        for lec in lectures:
            sql_pars = (lec['faculty_id'], lec['course_n'], lec['stream_n'], True)
            cur.execute(
                '''SELECT
                    vk_id
                FROM reports_student
                WHERE
                    faculty_id = %s AND
                    course_n = %s AND
                    stream_n = %s AND
                    vk_id IS NOT NULL AND
                    get_lec_urls_automatically = %s''',
                sql_pars
            )
            vk_ids = cur.fetchall()
            if not vk_ids:
                continue
            lec_dt = lec['lec_dt'].strftime('%d.%m в %H:%M')
            lec_url = f'Ссылка:\n{lec["lec_url"]}' if lec['lec_url'] != '' else 'Ссылки пока нет'
            msg = f'Лекция по {lec["subject"]} {lec_dt}\n{lec_url}'
            for user_set in seq_separator([str(x['vk_id']) for x in vk_ids], 99):
                var_message(user_ids=','.join(user_set), text=msg)
            cur.execute('UPDATE reports_lecturesschedule SET is_sent = %s WHERE id = %s', (True, lec['id']))
