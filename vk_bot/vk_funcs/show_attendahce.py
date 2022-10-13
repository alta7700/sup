from .config import *
from .all_buttons import *
from .excels_styles import *
from openpyxl import Workbook
from datetime import datetime, timedelta


def show_all_group_reports(user_id, stud_info):
    faculty_id = stud_info['faculty_id']
    course_n = stud_info['course_n']
    group_n = stud_info['group_n']

    cur.execute(
        '''SELECT
            ps.is_practical AS is_pr,
            s.short_title AS subject,
            ps.pr_lk_datetime AS pair_dt,
            ps.pair_n AS pair_n,
            mp.student_id AS st_id,
            mp.missing_reason AS reason,
            mp.is_debt AS is_debt
        FROM reports_pairsschedule AS ps
            JOIN reports_subject AS s
            ON s.id = ps.subject_id
            LEFT JOIN reports_missingpair AS mp
            ON mp.pair_id = ps.id
            LEFT JOIN (
                SELECT 
                    *
                FROM reports_student
                WHERE
                    faculty_id = %s AND
                    course_n = %s AND
                    group_n  = %s AND
                    is_fired = %s) AS st
            ON st.id = mp.student_id
        WHERE
            ps.faculty_id = %s AND
            ps.course_n = %s AND
            ps.group_n = %s AND
            ps.is_reported = %s
        ORDER BY
            ps.pr_lk_datetime,
            mp.student_id''',
        (faculty_id, course_n, group_n, False, faculty_id, course_n, group_n, True)
    )
    pairs = cur.fetchall()
    if not pairs:
        var_message(user_id, 'Рапортов ещё не было')
        return

    cur.execute(
        '''SELECT id, position_in_group AS pos, surname, name FROM reports_student
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty_id, course_n, group_n, False)
    )
    studs = {dict(x)['id']: dict(x) for x in cur.fetchall()}

    pairs_by_weeks = {}
    for pair in pairs:
        pair_info = (pair['pair_n'], pair['is_pr'], pair['subject'], pair['pair_dt'].replace(tzinfo=None))
        this_p_date = pair['pair_dt'].date()
        this_week_begin_dt = this_p_date - timedelta(days=this_p_date.weekday())
        if not pairs_by_weeks.get(this_week_begin_dt):
            pairs_by_weeks[this_week_begin_dt] = {}
        if not pairs_by_weeks[this_week_begin_dt].get(this_p_date):
            pairs_by_weeks[this_week_begin_dt][this_p_date] = {}
        if not pairs_by_weeks[this_week_begin_dt][this_p_date].get(pair_info):
            pairs_by_weeks[this_week_begin_dt][this_p_date][pair_info] = []
        if pair['st_id']:
            if studs.get(pair['st_id']):
                this_miss = (pair['st_id'], pair['reason'], pair['is_debt'])
                pairs_by_weeks[this_week_begin_dt][this_p_date][pair_info].append(this_miss)

    wb = Workbook()
    wb_sheet = wb.active
    br = 0
    for week, week_dates in pairs_by_weeks.items():
        start_this_date_col = -1
        extra_cols = 0
        wb_sheet.cell(row=br+1, column=1).value = '№'
        wb_sheet.cell(row=br+1, column=1).border = top_left_double_border
        wb_sheet.merge_cells(start_row=br+1, end_row=br+5, start_column=1, end_column=1)
        wb_sheet.column_dimensions['A'].width = 4
        wb_sheet.cell(row=br+1, column=2).value = 'Фамилия Имя'
        wb_sheet.cell(row=br+1, column=2).border = top_double_border
        wb_sheet.merge_cells(start_row=br+1, end_row=br+5, start_column=2, end_column=2)
        stud_names_len = []
        stud_id_to_row = {}
        for i, stud in enumerate(studs.values()):
            stud_id_to_row[stud['id']] = br+i+6
            wb_sheet.cell(row=br+i+6, column=1).value = stud['pos']
            wb_sheet.cell(row=br+i+6, column=1).border = left_double_border
            stud_name = f'{stud["surname"]} {stud["name"]}'
            wb_sheet.cell(row=br+i+6, column=2).value = stud_name
            wb_sheet.cell(row=br+i+6, column=2).border = thin_border
            stud_names_len.append(len(stud_name))
        wb_sheet.column_dimensions['B'].width = max(stud_names_len) * 1.3

        for week_date, pair_n in week_dates.items():
            start_this_date_col += 4 + extra_cols
            extra_cols = 0
            for p_info, missings in pair_n.items():
                if p_info[0] > 4:
                    extra_cols = max([extra_cols, p_info[0]]) - 4
                this_col = start_this_date_col + p_info[0] - 1
                wb_sheet.cell(row=br+3, column=this_col).value = p_info[0]
                wb_sheet.cell(row=br+3, column=this_col).alignment = Alignment(horizontal='center')

                wb_sheet.cell(row=br+4, column=this_col).value = 'пр' if p_info[1] else 'лк'
                wb_sheet.cell(row=br+4, column=this_col).alignment = Alignment(horizontal='center')

                wb_sheet.cell(row=br+5, column=this_col).value = p_info[2]
                wb_sheet.cell(row=br+5, column=this_col).alignment = Alignment(horizontal='center', textRotation=90)

                for i in range(len(studs)):
                    wb_sheet.cell(row=br+6+i, column=this_col).value = '+'
                    wb_sheet.cell(row=br+6+i, column=this_col).alignment = Alignment(horizontal='center')

                if missings:
                    for m in missings:
                        wb_sheet.cell(row=stud_id_to_row[m[0]], column=this_col).value = m[1]
                        if m[2]:
                            wb_sheet.cell(row=stud_id_to_row[m[0]], column=this_col).font = red_font

            for i in range(br+3, wb_sheet.max_row+1):
                wb_sheet.cell(row=i, column=start_this_date_col).border = left_double_border
                wb_sheet.cell(row=i, column=start_this_date_col+3+extra_cols).border = right_double_border
                for j in range(start_this_date_col+1, start_this_date_col+3+extra_cols):
                    wb_sheet.cell(row=i, column=j).border = thin_border

            wb_sheet.cell(row=br + 1, column=start_this_date_col).value = week_date.strftime('%A\n%d.%m.%Y')
            wb_sheet.cell(row=br + 1, column=start_this_date_col).border = all_double_border
            wb_sheet.cell(row=br + 1, column=start_this_date_col).alignment = date_alignment
            wb_sheet.merge_cells(start_row=br + 1, start_column=start_this_date_col,
                                 end_row=br + 2, end_column=start_this_date_col + 3 + extra_cols)
        max_row = wb_sheet.max_row+1
        for i in range(1, start_this_date_col+4+extra_cols):
            wb_sheet.cell(row=max_row, column=i).border = top_only_double_border

        br = wb_sheet.max_row + 4

    for i in range(3, wb_sheet.max_column+1):
        wb_sheet.column_dimensions[wb_sheet.cell(row=3, column=i).column_letter].width = 4

    wb_title = f'Все рапорты {get_faculty_by_id(faculty_id)["short_title"]} {course_n} курс {group_n} группа.xlsx'
    wb.save(wb_title)
    doc_sender(user_id, wb_title)


def show_reported_subjects_to_show_attendance(user_id, stud_info, is_practical):
    faculty_id = stud_info['faculty_id']
    course_n = stud_info['course_n']
    group_n = stud_info['group_n']
    cur.execute(
        '''SELECT
            s.id as id,
            MIN(ps.pr_lk_datetime) as min_time,
            s.short_title as short_title
        FROM
            reports_pairsschedule as ps
            LEFT JOIN reports_subject as s
            ON s.id = ps.subject_id
        WHERE
            ps.faculty_id = %s AND
            ps.course_n = %s AND
            ps.group_n = %s AND
            ps.is_reported = %s AND
            ps.is_practical = %s
        GROUP BY 
            s.id
        ORDER BY
            min_time''',
        (faculty_id, course_n, group_n, True, is_practical)
    )
    subjects = cur.fetchall()
    if not subjects:
        var_message(
            user_id, f'Рапортов по {"практическим занятиям" if is_practical else "лекциям"} еще нет',
            kb_params=main_kb_head if stud_info["is_head"] or stud_info['is_head_assistant'] else main_kb_student)
        return 0
    num = 2
    if stud_info['is_head'] or stud_info['is_head_assistant']:
        btn_payload_pattern = '{{{{"button":"attendance_show_{}-{{}}"}}}}'
        btn_payload = btn_payload_pattern.format('practical' if is_practical else 'lecture')
    else:
        btn_payload_pattern = '{{{{"button":"attendance_show_{}-{{}}-{}"}}}}'
        btn_payload = btn_payload_pattern.format('practical' if is_practical else 'lecture', stud_info['id'])
    subjects_kb = {1: [('Все предметы', 'positive', btn_payload.format('all')), ], num: []}
    max_btn_in_line = 3 if len(subjects) > 16 else 2
    for subj in subjects:
        subjects_kb[num].append((f'{subj["short_title"]}', 'primary', btn_payload.format(subj['id'])))
        if len(subjects_kb[num]) == max_btn_in_line:
            num += 1
            subjects_kb[num] = []
    subjects_kb[num].append(cancel_btn)
    var_message(user_id, '...', kb_params=subjects_kb)


def show_attendance_stud_choice_btn(user_id, head_info, payload):
    faculty_id = head_info['faculty_id']
    course_n = head_info['course_n']
    group_n = head_info['group_n']
    cur.execute(
        '''SELECT id as id, position_in_group as pos, surname as surname
        FROM reports_student
        WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_fired = %s
        ORDER BY id''',
        (faculty_id, course_n, group_n, False)
    )
    group_stud = cur.fetchall()
    btn_payload = f'{{{payload[:-2]}-{{}}"}}}}'
    num = 2
    group_kb = {1: [('Выбрать всех', 'positive', btn_payload.format('all')), ], num: []}
    max_btn_in_line = 3 if len(group_stud) > 17 else 2
    for stud in group_stud:
        group_kb[num].append((f'{stud["pos"]}. {stud["surname"]}', 'primary', btn_payload.format(stud['id'])))
        if len(group_kb[num]) == max_btn_in_line:
            num += 1
            group_kb[num] = []
    group_kb[num].append(cancel_btn)
    var_message(user_id, '...', kb_params=group_kb)


def show_attendance(user_id, stud_info, payload, is_practical):
    faculty_id = stud_info['faculty_id']
    course_n = stud_info['course_n']
    group_n = stud_info['group_n']

    sql_request = '''SELECT
            ps.course_n as course_n,
            ps.group_n as group_n,
            s.short_title as subject,
            ps.lecture_id as lecture_id,
            ps.pr_lk_datetime as pr_lk_datetime,
            mp_st.st_id as student_id,
            mp_st.pos as pos,
            mp_st.surname as surname,
            mp_st.name as name,
            mp_st.reason as reason,
            mp_st.is_debt as is_debt
        FROM reports_pairsschedule AS ps
            JOIN reports_subject as s
            ON s.id = ps.subject_id
            LEFT JOIN (
                SELECT
                    st.id as st_id,
                    st.is_fired as is_fired,
                    st.position_in_group as pos,
                    st.surname,
                    st.name,
                    mp.pair_id,
                    mp.missing_reason as reason,
                    mp.is_debt as is_debt
                FROM reports_missingpair as mp
                    JOIN reports_student as st
                    ON st.id = mp.student_id
                {}) as mp_st
            ON mp_st.pair_id = ps.id
        WHERE
            ps.faculty_id = %s AND
            ps.course_n = %s AND
            ps.group_n = %s AND
            ps.is_reported = %s AND
            ps.is_practical = %s'''
    sql_parameters = [faculty_id, course_n, group_n, True, is_practical]

    subj_id = payload[:-2].split('-')[1]
    if subj_id == 'all':
        all_subjects = True
    else:
        all_subjects = False
        subj_id = int(subj_id)
        sql_request += ' AND s.id = %s '
        sql_parameters.append(subj_id)

    stud_id = payload[:-2].split('-')[2]
    if stud_id == 'all':
        all_students = True
        sql_request = sql_request.format(
            'WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND st.is_fired = %s'
        )
        sql_parameters = [faculty_id, course_n, group_n, False] + sql_parameters
    else:
        all_students = False
        sql_request = sql_request.format('WHERE st.id = %s')
        stud_id = int(stud_id)
        sql_parameters.insert(0, stud_id)

    if not all_students and not all_subjects:
        sql_request += ' ORDER BY ps.pr_lk_datetime'
        cur.execute(sql_request, sql_parameters)
        pairs = cur.fetchall()
        cur.execute('SELECT surname, name FROM reports_student WHERE id = %s', (stud_id, ))
        this_stud = cur.fetchone()
        this_subj = pairs[0]["subject"]
        result_msg = f'{this_stud["surname"]} {this_stud["name"]}\n{this_subj}'
        for pair in pairs:
            pair_date = pair['pr_lk_datetime'].date().strftime('%d.%m')
            reason = pair['reason']
            is_debt = ' (отработано)' if not pair['is_debt'] else ' (не отработано)'
            if not reason:
                reason = '+'
            elif reason == 'н':
                reason = 'нб' + is_debt
            elif reason == 'б':
                reason = 'болел' + is_debt
            elif reason == 'с':
                reason = 'справка'
            result_msg += f'\n{pair_date} - {reason}'
        var_message(user_id, result_msg)

    elif not all_students and all_subjects:
        sql_request += ' ORDER BY s.id, ps.pr_lk_datetime'
        cur.execute(sql_request, sql_parameters)
        pairs = cur.fetchall()
        cur.execute('SELECT surname, name FROM reports_student WHERE id = %s', (stud_id, ))
        stud = cur.fetchone()
        wb = Workbook()
        wb_sheet = wb.active
        this_subj_col = -3
        this_row = 3
        last_subj = 'Такого точно не будет'
        for pair in pairs:
            this_subj = pair['subject']
            pair_dt = pair['pr_lk_datetime'].strftime('%d.%m')
            if this_subj != last_subj:
                this_subj_col += 4
                this_row = 3
                wb_sheet.cell(row=1, column=this_subj_col).value = this_subj
                wb_sheet.cell(row=1, column=this_subj_col).border = thin_border
                wb_sheet.cell(row=1, column=this_subj_col).alignment = center_cell
                wb_sheet.merge_cells(start_row=1, end_row=1, start_column=this_subj_col, end_column=this_subj_col+2)
                wb_sheet.cell(row=2, column=this_subj_col).value = 'Дата'
                wb_sheet.cell(row=2, column=this_subj_col).border = thin_border
                wb_sheet.cell(row=2, column=this_subj_col+1).value = '+/н/б/с'
                wb_sheet.cell(row=2, column=this_subj_col+1).border = thin_border
                wb_sheet.cell(row=2, column=this_subj_col+2).value = 'отработано'
                wb_sheet.cell(row=2, column=this_subj_col+2).border = thin_border
                wb_sheet.column_dimensions[wb_sheet.cell(row=2, column=this_subj_col+2).column_letter].wigth = 12
            for col in wb_sheet.columns:
                if col[1].value:
                    for cell in col:
                        cell.alignment = center_cell

            reason = pair['reason']
            is_debt = '+' if not pair['is_debt'] else '-'
            if not reason:
                reason = '+'
            elif reason == 'н':
                reason = 'нб'
                wb_sheet.cell(row=this_row, column=this_subj_col + 2).value = is_debt
            elif reason == 'б':
                reason = 'болел'
                wb_sheet.cell(row=this_row, column=this_subj_col + 2).value = is_debt
            elif reason == 'с':
                reason = 'справка'
            wb_sheet.cell(row=this_row, column=this_subj_col).value = pair_dt
            wb_sheet.cell(row=this_row, column=this_subj_col).border = thin_border
            wb_sheet.cell(row=this_row, column=this_subj_col).alignment = center_cell
            wb_sheet.cell(row=this_row, column=this_subj_col+1).value = reason
            wb_sheet.cell(row=this_row, column=this_subj_col+1).border = thin_border
            wb_sheet.cell(row=this_row, column=this_subj_col+2).border = thin_border
            last_subj = this_subj
            this_row += 1
        faculty = get_faculty_by_id(faculty_id)['short_title']
        wb_title = f'Посещаемость {faculty} {course_n} курс {group_n} группа {stud["surname"]} {stud["name"]}.xlsx'
        wb.save(wb_title)
        kb = main_kb_student if not stud_info['is_head'] and not stud_info['is_head_assistant'] else None
        doc_sender(user_id, wb_title, kb_params=kb)

    elif all_students and not all_subjects:
        sql_request += ' ORDER BY ps.pr_lk_datetime, mp_st.st_id'
        cur.execute(sql_request, sql_parameters)
        pairs = cur.fetchall()
        wb = Workbook()
        wb_sheet = wb.active
        this_subj = pairs[0]['subject']

        cur.execute(
            '''SELECT id, position_in_group AS pos, surname, name FROM reports_student
            WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_fired = %s
            ORDER BY id''',
            (faculty_id, course_n, group_n, False)
        )
        group_stud = cur.fetchall()
        make_multi_student_table(wb_sheet, group_stud, this_subj, pairs, 0)

        faculty = get_faculty_by_id(faculty_id)['short_title']
        wb_title = f'Посещаемость {faculty} {course_n} курс {group_n} группа {this_subj}.xlsx'
        wb.save(wb_title)
        doc_sender(user_id, wb_title,
                   kb_params=main_kb_student if not stud_info['is_head'] and not stud_info['is_head_assistant'] else None)

    elif all_students and all_subjects:
        sql_request += ' ORDER BY s.id, ps.pr_lk_datetime, mp_st.st_id'
        cur.execute(sql_request, sql_parameters)
        pairs = cur.fetchall()
        cur.execute(
            '''SELECT id, position_in_group AS pos, surname, name FROM reports_student
            WHERE faculty_id = %s AND course_n = %s AND group_n = %s AND is_fired = %s
            ORDER BY id''',
            (faculty_id, course_n, group_n, False)
        )
        group_stud = cur.fetchall()

        wb = Workbook()
        wb_sheet = wb.active
        last_subject = pairs[0]['subject']
        such_subjects = []
        pairs_dict = {}
        for pair in pairs:
            this_subject = pair['subject']
            if this_subject == last_subject:
                such_subjects.append(dict(pair))
            else:
                pairs_dict[last_subject] = such_subjects
                such_subjects = [(dict(pair)), ]
            last_subject = this_subject
        pairs_dict[last_subject] = such_subjects
        begin_row = 0
        for subject, pairs in pairs_dict.items():
            this_subject = pairs[0]['subject']
            make_multi_student_table(wb_sheet, group_stud, this_subject, pairs, begin_row)
            begin_row = wb_sheet.max_row + 3
        faculty = get_faculty_by_id(faculty_id)['short_title']
        wb_title = f'Посещаемость {faculty} {course_n} курс {group_n} группа.xlsx'
        wb.save(wb_title)
        doc_sender(user_id, wb_title)


def make_multi_student_table(wb_sheet, group_stud, this_subj, pairs, br):
    stud_id_to_row = {}
    len_studs = []
    row_to_nb = {}
    for i, stud in enumerate(group_stud):
        stud_id_to_row[stud['id']] = br + i + 4
        row_to_nb[br + i + 4] = {'н': 0, 'б': 0, 'с': 0, 'долг': 0}
        stud_name = f'{stud["surname"]} {stud["name"]}'
        wb_sheet.cell(row=br+i+4, column=1).value = stud['pos']
        wb_sheet.cell(row=br+i+4, column=1).border = left_double_border
        wb_sheet.cell(row=br+i+4, column=2).value = stud_name
        wb_sheet.cell(row=br+i+4, column=2).border = thin_border
        len_studs.append(len(stud_name))
    wb_sheet.column_dimensions['A'].width = 4
    wb_sheet.column_dimensions['B'].width = max(len_studs) * 1.3

    last_pair_date = 'Какая-то дата'
    this_pair_col = 1
    wb_sheet.cell(row=br+1, column=1).value = '№'
    wb_sheet.cell(row=br+1, column=1).border = top_left_double_border
    wb_sheet.merge_cells(start_row=br+1, end_row=br+3, start_column=1, end_column=1)
    wb_sheet.cell(row=br+1, column=2).value = 'Фамилия Имя'
    wb_sheet.cell(row=br+1, column=2).border = top_double_border
    wb_sheet.merge_cells(start_row=br+1, end_row=br+3, start_column=2, end_column=2)
    wb_sheet.cell(row=br+1, column=3).value = this_subj
    wb_sheet.cell(row=br+1, column=3).border = all_double_border
    wb_sheet.cell(row=br+1, column=3).alignment = center_cell
    for i, pair in enumerate(pairs):
        pair_date = pair['pr_lk_datetime'].strftime('%d.%m %H:%M')
        if last_pair_date != pair_date:
            this_pair_col += 2
            wb_sheet.cell(row=br+2, column=this_pair_col).value = pair_date
            wb_sheet.cell(row=br+2, column=this_pair_col).border = left_right_double_border
            wb_sheet.cell(row=br+2, column=this_pair_col).alignment = center_cell
            wb_sheet.merge_cells(start_row=br+2, end_row=br+2, start_column=this_pair_col, end_column=this_pair_col+1)

            reason_cell = wb_sheet.cell(row=br+3, column=this_pair_col)
            wb_sheet.column_dimensions[reason_cell.column_letter].width = 5
            reason_cell.value = '+/-'
            reason_cell.border = left_double_border
            reason_cell.alignment = center_cell

            debt_cell = wb_sheet.cell(row=br+3, column=this_pair_col+1)
            wb_sheet.column_dimensions[debt_cell.column_letter].width = 6
            debt_cell.value = 'отр.'
            debt_cell.border = right_double_border
            debt_cell.alignment = center_cell
            ii = 4
            while wb_sheet.cell(row=br+ii, column=1).value:
                wb_sheet.cell(row=br+ii, column=this_pair_col).value = '+'
                wb_sheet.cell(row=br+ii, column=this_pair_col).border = left_double_border
                wb_sheet.cell(row=br+ii, column=this_pair_col).alignment = center_cell
                wb_sheet.cell(row=br+ii, column=this_pair_col+1).border = right_double_border
                wb_sheet.cell(row=br+ii, column=this_pair_col+1).alignment = center_cell
                ii += 1
        if pair['student_id']:
            row_of_stud = stud_id_to_row[pair['student_id']]
            reason = pair['reason']
            if reason == 'с':
                reason = 'спр'
                row_to_nb[row_of_stud]['с'] = row_to_nb[row_of_stud]['с'] + 1
            elif reason == 'н':
                reason = 'нб'
                row_to_nb[row_of_stud]['н'] = row_to_nb[row_of_stud]['н'] + 1
            elif reason == 'б':
                reason = 'бол'
                row_to_nb[row_of_stud]['б'] = row_to_nb[row_of_stud]['б'] + 1
            wb_sheet.cell(row=row_of_stud, column=this_pair_col).value = reason
            if reason != 'спр':
                if not pair['is_debt']:
                    wb_sheet.cell(row=row_of_stud, column=this_pair_col+1).value = '+'
                else:
                    wb_sheet.cell(row=row_of_stud, column=this_pair_col + 1).value = '-'
                    wb_sheet.cell(row=row_of_stud, column=this_pair_col+1).font = red_font
                    row_to_nb[row_of_stud]['долг'] = row_to_nb[row_of_stud]['долг'] + 1
        last_pair_date = pair_date

    wb_sheet.cell(row=br+2, column=this_pair_col + 2).value = 'Без уваж'
    wb_sheet.cell(row=br+2, column=this_pair_col + 3).value = 'Болел'
    wb_sheet.cell(row=br+2, column=this_pair_col + 4).value = 'Справка'
    wb_sheet.cell(row=br+2, column=this_pair_col + 5).value = 'Общее'
    wb_sheet.cell(row=br+2, column=this_pair_col + 6).value = 'Долг'

    for i in range(2, 7):
        if i != 2:
            wb_sheet.cell(row=br+2, column=this_pair_col + i).border = thin_border
        else:
            wb_sheet.cell(row=br+2, column=this_pair_col + i).border = left_double_border
        wb_sheet.cell(row=br+2, column=this_pair_col + i).alignment = center_cell
        wb_sheet.merge_cells(start_row=br+2, end_row=br+3, start_column=this_pair_col+i, end_column=this_pair_col+i)

    for r, count in row_to_nb.items():
        wb_sheet.cell(row=r, column=this_pair_col + 2).value = row_to_nb[r]['н']
        wb_sheet.cell(row=r, column=this_pair_col + 2).border = left_double_border
        wb_sheet.cell(row=r, column=this_pair_col + 2).alignment = center_cell

        wb_sheet.cell(row=r, column=this_pair_col + 3).value = count['б']
        wb_sheet.cell(row=r, column=this_pair_col + 4).value = count['с']
        wb_sheet.cell(row=r, column=this_pair_col + 5).value = count['н'] + count['б'] + count['с']
        wb_sheet.cell(row=r, column=this_pair_col + 6).value = count['долг']
        wb_sheet.cell(row=r, column=this_pair_col + 6).font = red_font
        for i in range(3, 7):
            wb_sheet.cell(row=r, column=this_pair_col + i).border = thin_border
            wb_sheet.cell(row=r, column=this_pair_col + i).alignment = center_cell
    for i in range(br+1, wb_sheet.max_row):
        wb_sheet.cell(row=i, column=this_pair_col+7).border = left_only_double_border
    for i in range(1, this_pair_col + 7):
        max_row = wb_sheet.max_row
        wb_sheet.cell(row=max_row, column=i).border = top_only_double_border
    wb_sheet.merge_cells(start_row=br+1, end_row=br+1, start_column=3, end_column=this_pair_col + 6)


# def show_subject_to_add_pair_comment(user_id, stud_info):
#     faculty_id = stud_info['faculty_id']
#     course_n = stud_info['course_n']
#     group_n = stud_info['group_n']
#
#
#
# def show_dates_to_add_pair_comment():
#     pass
#
#
# def send_comment_add_pattern():
#     pass
