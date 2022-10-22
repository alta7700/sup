from openpyxl import Workbook
from .excels_styles import *
from .models import *
from django.db.models.functions.datetime import datetime


def show_attendance_group(faculty_id, course_n, group_n, stud_id='all', subj_id='all', is_pr=True):
    all_subjects = True if subj_id == 'all' else False
    all_students = True if stud_id == 'all' else False

    if not all_students and not all_subjects:
        this_stud = Student.objects.values('surname', 'name').filter(id=stud_id, faculty=faculty_id, course_n=course_n,
                                                                     group_n=group_n)
        if not this_stud:
            return 'Не нужно менять форму'
        this_stud = this_stud[0]
        pairs = PairsSchedule.objects.raw(
            '''SELECT
                ps.id,
                s.short_title,
                ps.pr_lk_datetime,
                mp.missing_reason,
                mp.is_debt
            FROM reports_pairsschedule AS ps
                JOIN reports_subject as s
                ON s.id = ps.subject_id
                LEFT JOIN (SELECT * FROM reports_missingpair WHERE student_id = %s) as mp   
                ON mp.pair_id = ps.id
            WHERE
                ps.faculty_id = %s AND
                ps.course_n = %s AND
                ps.group_n = %s AND
                s.id = %s AND
                ps.is_reported = %s AND
                ps.is_practical = %s
            ORDER BY
                ps.pr_lk_datetime''',
            (stud_id, faculty_id, course_n, group_n, subj_id, True, is_pr)
        )
        this_subj = Subject.objects.get(id=subj_id).short_title
        if not pairs:
            pairs = PairsSchedule.objects.filter(
                faculty__id=faculty_id, course_n=course_n, group_n=group_n, subject__id=subj_id,
                is_practical=is_pr, pr_lk_datetime__lt=datetime.now())
            if pairs:
                return f'Рапортичек за {"практические занятия" if is_pr else "лекции"} по {this_subj} ещё нет'
            else:
                return f'Ни {"одного практического занятия" if is_pr else "одной лекции"} по {this_subj} ещё не было'

        result_msg = f'{this_stud["surname"]} {this_stud["name"]}\n{this_subj}'
        for pair in pairs:
            pair_date = pair.pr_lk_datetime.date().strftime('%d.%m')
            reason = pair.missing_reason
            if not reason:
                reason = '+'
            elif reason == 'н':
                reason = f"'нб' {'(отработано)' if not pair.is_debt else '(не отработано)'}"
            elif reason == 'б':
                reason = f"'болел' {'(отработано)' if not pair.is_debt else '(не отработано)'}"
            elif reason == 'с':
                reason = 'справка'
            result_msg += f'\n{pair_date} - {reason}'
        return result_msg

    elif not all_students and all_subjects:
        stud = Student.objects.values('surname', 'name').filter(id=stud_id, faculty=faculty_id, course_n=course_n,
                                                                     group_n=group_n)
        if not stud:
            return 'Не нужно менять форму'
        stud = stud[0]
        pairs = PairsSchedule.objects.raw(
            '''SELECT
                ps.id,
                ps.pr_lk_datetime,
                s.short_title,
                mp.missing_reason,
                mp.is_debt
            FROM reports_pairsschedule AS ps
                JOIN reports_subject as s
                ON s.id = ps.subject_id
                LEFT JOIN (SELECT * FROM reports_missingpair WHERE student_id = %s) as mp   
                ON mp.pair_id = ps.id
            WHERE
                ps.faculty_id = %s AND
                ps.course_n = %s AND
                ps.group_n = %s AND
                ps.is_reported = %s AND
                ps.is_practical = %s
            ORDER BY
                s.id, 
                ps.pr_lk_datetime''',
            (stud_id, faculty_id, course_n, group_n, True, is_pr)
        )

        if not pairs:
            pairs = PairsSchedule.objects.filter(faculty__id=faculty_id, course_n=course_n, group_n=group_n,
                                                 is_practical=is_pr, pr_lk_datetime__lt=datetime.now())
            if pairs:
                msg = f'Рапортичек по {"практическим занятиям" if is_pr else "лекциям"} ещё нет'
            else:
                msg = f'Ни {"одного практического занятия" if is_pr else "одной лекции"} ещё не было'
            return {'e_msg': msg}
        pairs_by_subject = {}
        for pair in pairs:
            s = pair.short_title
            if not pairs_by_subject.get(s):
                pairs_by_subject[s] = []
            pairs_by_subject[s].append(pair)

        wb = Workbook()
        wb_sheet = wb.active
        bc = 0
        for subject, pairs in pairs_by_subject.items():
            wb_sheet.cell(row=1, column=bc+1).value = subject
            wb_sheet.cell(row=1, column=bc+1).alignment = center_cell
            wb_sheet.cell(row=2, column=bc+1).value = 'Дата'
            wb_sheet.cell(row=2, column=bc+1).border = left_double_border
            wb_sheet.cell(row=2, column=bc+2).value = '+/н/б/с'
            wb_sheet.cell(row=2, column=bc+2).border = thin_border
            wb_sheet.cell(row=2, column=bc+3).value = 'отраб.'
            wb_sheet.cell(row=2, column=bc+3).border = right_double_border
            wb_sheet.cell(row=1, column=bc+1).border = all_double_border
            wb_sheet.merge_cells(start_row=1, end_row=1, start_column=bc+1, end_column=bc+3)
            this_row = 3
            debts = 0
            for pair in pairs:
                pair_dt = pair.pr_lk_datetime.strftime('%d.%m')
                reason = pair.missing_reason
                is_debt = '+' if not pair.is_debt else '-'
                if not reason:
                    reason = '+'
                elif reason == 'н':
                    reason = 'нб'
                    wb_sheet.cell(row=this_row, column=bc+3).value = is_debt
                    if is_debt == '-':
                        debts += 1
                        wb_sheet.cell(row=this_row, column=bc + 3).font = red_font
                elif reason == 'б':
                    reason = 'болел'
                    wb_sheet.cell(row=this_row, column=bc+3).value = is_debt
                    if is_debt == '-':
                        debts += 1
                        wb_sheet.cell(row=this_row, column=bc+3).font = red_font
                elif reason == 'с':
                    reason = 'справка'
                wb_sheet.cell(row=this_row, column=bc+1).value = pair_dt
                wb_sheet.cell(row=this_row, column=bc+1).border = thin_border
                wb_sheet.cell(row=this_row, column=bc+2).value = reason
                this_row += 1

            for i in range(3, this_row):
                wb_sheet.cell(row=i, column=bc+1).border = left_double_border
                wb_sheet.cell(row=i, column=bc+2).border = thin_border
                wb_sheet.cell(row=i, column=bc+3).border = right_double_border
            for i in range(bc+1, bc+4):
                wb_sheet.cell(row=this_row, column=i).border = top_only_double_border

            for i in range(2, this_row):
                for j in range(bc+1, bc+5):
                    wb_sheet.cell(row=i, column=j).alignment = center_cell
            bc += 4
        faculty = Faculty.objects.values('short_title').get(id=faculty_id)['short_title']
        wb_title = f'Посещаемость {faculty} {course_n} курс {group_n} группа {stud["surname"]} {stud["name"]}.xlsx'
        return {'workbook': wb, 'filename': wb_title}

    elif all_students and not all_subjects:
        pairs = PairsSchedule.objects.raw(
            '''SELECT
                ps.id,
                ps.pr_lk_datetime,
                s.short_title,
                mp.student_id,
                mp.missing_reason,
                mp.is_debt
            FROM reports_pairsschedule AS ps
                JOIN reports_subject as s
                ON s.id = ps.subject_id
                LEFT JOIN reports_missingpair as mp   
                ON mp.pair_id = ps.id
            WHERE
                ps.faculty_id = %s AND
                ps.course_n = %s AND
                ps.group_n = %s AND
                s.id = %s AND
                ps.is_reported = %s AND
                ps.is_practical = %s
            ORDER BY
                ps.pr_lk_datetime,
                mp.student_id''',
            (faculty_id, course_n, group_n, subj_id, True, is_pr)
        )
        this_subj = Subject.objects.get(id=subj_id).short_title
        if not pairs:
            pairs = PairsSchedule.objects.filter(faculty__id=faculty_id, course_n=course_n, group_n=group_n,
                                                 subject=subj_id, is_practical=is_pr, pr_lk_datetime__lt=datetime.now())
            if pairs:
                msg = f'Рапортичек за {"практические занятия" if is_pr else "лекции"} по {this_subj} ещё нет'
            else:
                msg = f'Ни {"одного практического занятия" if is_pr else "одной лекции"} по {this_subj} ещё не было'
            return {'e_msg': msg}

        group_studs = Student.objects.filter(faculty__id=faculty_id, course_n=course_n, group_n=group_n, is_fired=False)
        wb = Workbook()
        wb_sheet = wb.active

        make_group_subject_table(wb_sheet, group_studs, this_subj, pairs, 0)
        faculty = Faculty.objects.values('short_title').get(id=faculty_id)['short_title']
        wb_title = f'Посещаемость {faculty} {course_n} курс {group_n} группа {this_subj}.xlsx'
        return {'workbook': wb, 'filename': wb_title}

    elif all_students and all_subjects:
        pairs = PairsSchedule.objects.raw(
            '''SELECT
                ps.id,
                ps.pr_lk_datetime,
                s.short_title,
                mp.student_id,
                mp.missing_reason,
                mp.is_debt
            FROM reports_pairsschedule AS ps
                JOIN reports_subject as s
                ON s.id = ps.subject_id
                LEFT JOIN reports_missingpair as mp   
                ON mp.pair_id = ps.id
            WHERE
                ps.faculty_id = %s AND
                ps.course_n = %s AND
                ps.group_n = %s AND
                ps.is_reported = %s AND
                ps.is_practical = %s
            ORDER BY
                s.id,
                ps.pr_lk_datetime,
                mp.student_id''',
            (faculty_id, course_n, group_n, True, is_pr)
        )
        if not pairs:
            pairs = PairsSchedule.objects.filter(faculty__id=faculty_id, course_n=course_n, group_n=group_n,
                                                 is_practical=is_pr, pr_lk_datetime__lt=datetime.now())
            if pairs:
                msg = f'Рапортичек за {"практические занятия" if is_pr else "лекции"} ещё нет'
            else:
                msg = f'Ни {"одного практического занятия" if is_pr else "одной лекции"} ещё не было'
            return {'e_msg': msg}

        pairs_by_subject = {}
        for pair in pairs:
            if not pairs_by_subject.get(pair.short_title):
                pairs_by_subject[pair.short_title] = []
            pairs_by_subject[pair.short_title].append(pair)

        group_studs = Student.objects.filter(faculty__id=faculty_id, course_n=course_n, group_n=group_n, is_fired=False)
        wb = Workbook()
        wb_sheet = wb.active
        br = 0
        for subj_title, subj_pairs in pairs_by_subject.items():
            make_group_subject_table(wb_sheet, group_studs, subj_title, subj_pairs, br)
            br = wb_sheet.max_row + 3
        faculty = Faculty.objects.values('short_title').get(id=faculty_id)['short_title']
        wb_title = f'Посещаемость {faculty} {course_n} курс {group_n} группа.xlsx'
        return {'workbook': wb, 'filename': wb_title}


def make_group_subject_table(wb_sheet, group_stud, this_subj, pairs, br):
    pairs_by_dt = {}
    for pair in pairs:
        d = pair.pr_lk_datetime
        if not pairs_by_dt.get(d):
            pairs_by_dt[d] = []
        if pair.student_id:
            pairs_by_dt[d].append(pair)

    stud_id_to_row = {}
    len_studs = []
    row_to_nb = {}
    for i, stud in enumerate(group_stud):
        stud_id_to_row[stud.id] = br + i + 4
        row_to_nb[br+i+4] = {'н': 0, 'б': 0, 'с': 0, 'долг': 0}
        stud_name = f'{stud.surname} {stud.name}'
        wb_sheet.cell(row=br+i+4, column=1).value = stud.position_in_group
        wb_sheet.cell(row=br+i+4, column=1).border = left_double_border
        wb_sheet.cell(row=br+i+4, column=2).value = stud_name
        wb_sheet.cell(row=br+i+4, column=2).border = thin_border
        len_studs.append(len(stud_name))
    wb_sheet.column_dimensions['A'].width = 4
    wb_sheet.column_dimensions['B'].width = max(len_studs) * 1.3

    this_pair_col = 3
    wb_sheet.cell(row=br+1, column=1).value = '№'
    wb_sheet.cell(row=br+1, column=1).border = top_left_double_border
    wb_sheet.merge_cells(start_row=br+1, end_row=br+3, start_column=1, end_column=1)
    wb_sheet.cell(row=br+1, column=2).value = 'Фамилия Имя'
    wb_sheet.cell(row=br+1, column=2).border = top_double_border
    wb_sheet.merge_cells(start_row=br+1, end_row=br+3, start_column=2, end_column=2)
    wb_sheet.cell(row=br+1, column=3).value = this_subj
    wb_sheet.cell(row=br+1, column=3).border = all_double_border
    wb_sheet.cell(row=br+1, column=3).alignment = center_cell
    for pair_datetime, one_dt_pairs in pairs_by_dt.items():
        pair_dt = pair_datetime.strftime('%d.%m %H:%M')
        pair_dt_cell = wb_sheet.cell(row=br+2, column=this_pair_col)
        pair_dt_cell.value = pair_dt
        pair_dt_cell.border = left_right_double_border
        pair_dt_cell.alignment = center_cell
        wb_sheet.merge_cells(start_row=br+2, end_row=br+2, start_column=this_pair_col, end_column=this_pair_col+1)

        reason_cell = wb_sheet.cell(row=br+3, column=this_pair_col)
        wb_sheet.column_dimensions[reason_cell.column_letter].width = 5
        reason_cell.value = '+/-'
        reason_cell.border = left_double_border
        reason_cell.alignment = center_cell

        debt_cell = wb_sheet.cell(row=br+3, column=this_pair_col+1)
        wb_sheet.column_dimensions[debt_cell.column_letter].width = 6
        debt_cell.value = 'отр.'
        debt_cell.border = right_double_border
        debt_cell.alignment = center_cell
        for i in range(4, len(group_stud)+4):
            wb_sheet.cell(row=br+i, column=this_pair_col).value = '+'
            wb_sheet.cell(row=br+i, column=this_pair_col).border = left_double_border
            wb_sheet.cell(row=br+i, column=this_pair_col).alignment = center_cell
            wb_sheet.cell(row=br+i, column=this_pair_col+1).border = right_double_border
            wb_sheet.cell(row=br+i, column=this_pair_col+1).alignment = center_cell
        for pair in one_dt_pairs:
            row_of_stud = stud_id_to_row.get(pair.student_id)
            if row_of_stud:
                reason = pair.missing_reason
                if reason == 'с':
                    reason = 'спр'
                    row_to_nb[row_of_stud]['с'] = row_to_nb[row_of_stud]['с'] + 1
                elif reason == 'н':
                    reason = 'нб'
                    row_to_nb[row_of_stud]['н'] = row_to_nb[row_of_stud]['н'] + 1
                elif reason == 'б':
                    reason = 'бол'
                    row_to_nb[row_of_stud]['б'] = row_to_nb[row_of_stud]['б'] + 1
                wb_sheet.cell(row=row_of_stud, column=this_pair_col).value = reason
                if reason != 'спр':
                    if not pair.is_debt:
                        wb_sheet.cell(row=row_of_stud, column=this_pair_col+1).value = '+'
                    else:
                        wb_sheet.cell(row=row_of_stud, column=this_pair_col + 1).value = '-'
                        wb_sheet.cell(row=row_of_stud, column=this_pair_col+1).font = red_font
                        row_to_nb[row_of_stud]['долг'] = row_to_nb[row_of_stud]['долг'] + 1
        this_pair_col += 2

    wb_sheet.cell(row=br+2, column=this_pair_col).value = 'Без уваж'
    wb_sheet.cell(row=br+2, column=this_pair_col+1).value = 'Болел'
    wb_sheet.cell(row=br+2, column=this_pair_col+2).value = 'Справка'
    wb_sheet.cell(row=br+2, column=this_pair_col+3).value = 'Общее'
    wb_sheet.cell(row=br+2, column=this_pair_col+4).value = 'Долг'
    for i in range(5):
        if i != 0:
            wb_sheet.cell(row=br+2, column=this_pair_col+i).border = thin_border
        else:
            wb_sheet.cell(row=br+2, column=this_pair_col+i).border = left_double_border
        wb_sheet.cell(row=br+2, column=this_pair_col+i).alignment = center_cell
        wb_sheet.merge_cells(start_row=br+2, end_row=br+3, start_column=this_pair_col+i, end_column=this_pair_col+i)

    for r, count in row_to_nb.items():
        wb_sheet.cell(row=r, column=this_pair_col).value = row_to_nb[r]['н']
        wb_sheet.cell(row=r, column=this_pair_col).border = left_double_border
        wb_sheet.cell(row=r, column=this_pair_col).alignment = center_cell

        wb_sheet.cell(row=r, column=this_pair_col+1).value = count['б']
        wb_sheet.cell(row=r, column=this_pair_col+2).value = count['с']
        wb_sheet.cell(row=r, column=this_pair_col+3).value = count['н'] + count['б'] + count['с']
        wb_sheet.cell(row=r, column=this_pair_col+4).value = count['долг']
        wb_sheet.cell(row=r, column=this_pair_col+4).font = red_font
        for i in range(1, 5):
            wb_sheet.cell(row=r, column=this_pair_col + i).border = thin_border
            wb_sheet.cell(row=r, column=this_pair_col + i).alignment = center_cell
    for i in range(br+1, wb_sheet.max_row+1):
        wb_sheet.cell(row=i, column=this_pair_col+5).border = left_only_double_border

    wb_sheet.merge_cells(start_row=br+1, end_row=br+1, start_column=3, end_column=this_pair_col+4)
    max_row = wb_sheet.max_row + 1
    for i in range(1, this_pair_col+5):
        wb_sheet.cell(row=max_row, column=i).border = top_only_double_border


def show_attendance_course(faculty_id, course_n, stream_n=None, subj_id='all', is_pr=True):
    all_subjects = True if subj_id == 'all' else False

    sql_request = '''SELECT
        ps.id,
        ps.group_n,
        ps.pr_lk_datetime,
        s.id AS s_id,
        s.short_title,
        mp.student_id,
        mp.missing_reason,
        mp.is_debt
    FROM reports_pairsschedule AS ps
        JOIN reports_subject AS s
        ON s.id = ps.subject_id
        LEFT JOIN reports_missingpair AS mp
        ON mp.pair_id = ps.id
    WHERE
        ps.faculty_id = %s AND
        ps.course_n = %s AND
        {{}}
        {}
        ps.is_reported = %s AND
        ps.is_practical = %s
    ORDER BY
        ps.group_n,
        s.id,
        ps.pr_lk_datetime 
        '''

    sql_pars = [faculty_id, course_n, True, is_pr]
    if not all_subjects:
        sql_request = sql_request.format('s.id = %s AND')
        sql_pars.insert(2, subj_id)
    else:
        sql_request = sql_request.format('')

    if stream_n:
        studs = list(Student.objects.filter(faculty=faculty_id, course_n=course_n, stream_n=stream_n, is_fired=False))
        if not studs:
            return {'e_msg': f'На этом курсе нет {stream_n} потока'}
        groups = []
        for stud in studs:
            if stud.group_n not in groups:
                groups.append(stud.group_n)
        sql_request = sql_request.format('ps.group_n = ANY(%s) AND')
        sql_pars.insert(2, groups)
    else:
        studs = list(Student.objects.filter(faculty=faculty_id, course_n=course_n, is_fired=False))
        if not studs:
            return {'e_msg': f'Нет списка этого курса'}
        sql_request = sql_request.format('')

    pairs = PairsSchedule.objects.raw(sql_request, sql_pars)
    if not pairs:
        if stream_n:
            pairs = PairsSchedule.objects.filter(faculty__id=faculty_id, course_n=course_n, group_n__in=groups,
                                                 is_practical=is_pr, pr_lk_datetime__lt=datetime.now())
        else:
            pairs = PairsSchedule.objects.filter(faculty__id=faculty_id, course_n=course_n,
                                                 is_practical=is_pr, pr_lk_datetime__lt=datetime.now())
        if not all_subjects:
            try:
                s_info = Subject.objects.get(id=subj_id)
                s_title = s_info.short_title
            except Subject.DoesNotExist:
                return {'e_msg': 'Какая-то ошибка, не получается обработать запрос'}
        if pairs:
            msg = f'Рапортичек за {"практические занятия" if is_pr else "лекции"}' \
                  f'{"" if all_subjects else f" {s_title}"} ещё нет'
        else:
            msg = f'Ни {"одного практического занятия" if is_pr else "одной лекции"}' \
                  f'{"" if all_subjects else f" {s_title}"} ещё не было'
        return {'e_msg': msg}

    def fill_header(bc, title):
        wb_sheet.cell(1, bc).value = title
        wb_sheet.cell(1, bc).border = all_double_border
        wb_sheet.cell(1, bc).alignment = center_cell
        wb_sheet.merge_cells(start_row=1, end_row=1, start_column=bc, end_column=bc+10)
        wb_sheet.cell(2, bc).value = 'пар'
        wb_sheet.cell(2, bc).border = left_double_border
        wb_sheet.cell(2, bc).alignment = center_cell
        wb_sheet.cell(3, bc).value = 'шт'
        wb_sheet.cell(2, bc+1).value = 'причина нб'
        wb_sheet.cell(2, bc+1).border = thin_border
        wb_sheet.cell(2, bc+1).alignment = center_cell
        wb_sheet.merge_cells(start_row=2, end_row=2, start_column=bc+1, end_column=bc+6)
        wb_sheet.cell(3, bc+1).value = 'н,шт'
        wb_sheet.cell(3, bc+2).value = 'н,%'
        wb_sheet.cell(3, bc+3).value = 'б,шт'
        wb_sheet.cell(3, bc+4).value = 'б,%'
        wb_sheet.cell(3, bc+5).value = 'с,шт'
        wb_sheet.cell(3, bc+6).value = 'с,%'
        wb_sheet.cell(2, bc+7).value = 'всего нб'
        wb_sheet.cell(2, bc+7).border = thin_border
        wb_sheet.cell(2, bc+7).alignment = center_cell
        wb_sheet.merge_cells(start_row=2, end_row=2, start_column=bc+7, end_column=bc+8)
        wb_sheet.cell(3, bc+7).value = 'шт'
        wb_sheet.cell(3, bc+8).value = '%'
        wb_sheet.cell(2, bc+9).value = 'неотраб.'
        wb_sheet.cell(2, bc+9).border = right_double_border
        wb_sheet.cell(2, bc+9).alignment = center_cell
        wb_sheet.merge_cells(start_row=2, end_row=2, start_column=bc+9, end_column=bc+10)
        wb_sheet.cell(3, bc+9).value = 'шт'
        wb_sheet.cell(3, bc+10).value = '%'

        for ii in range(3, wb_sheet.max_row+1):
            wb_sheet.cell(ii, bc).border = left_double_border
            wb_sheet.cell(ii, bc).alignment = center_cell
            wb_sheet.cell(ii, bc+10).border = right_double_border
            wb_sheet.cell(ii, bc+10).alignment = center_cell
            for col in range(1, 10):
                wb_sheet.cell(ii, bc+col).border = thin_border
                wb_sheet.cell(ii, bc+col).alignment = center_cell
        for ii in range(bc, bc+11):
            if (ii - bc) % 2 == 0 and ii != bc:
                wb_sheet.column_dimensions[wb_sheet.cell(3, ii).column_letter].width = 5.5
            else:
                wb_sheet.column_dimensions[wb_sheet.cell(3, ii).column_letter].width = 4.5

    def fill_col_with_percent(st_row, subj_col, p_count, n_count):
        wb_sheet.cell(st_row, subj_col).value = n_count
        percent_n = 100 * n_count // p_count
        wb_sheet.cell(st_row, subj_col+1).value = percent_n
        if percent_n > 30:
            wb_sheet.cell(st_row, subj_col + 1).font = red_font

    if all_subjects:
        all_subj = list(Subject.objects.filter(faculty=faculty_id, course_n=course_n).order_by('id'))

        studs_by_id = {stud.id: stud for stud in studs}
        studs_id_by_groups = {}
        for stud in studs:
            if not studs_id_by_groups.get(stud.group_n):
                studs_id_by_groups[stud.group_n] = {}
            studs_id_by_groups[stud.group_n][stud.id] =\
                {subj.id: {'пар': 0, 'н': 0, 'б': 0, 'с': 0, 'долг': 0} for subj in all_subj}

        wb = Workbook()
        wb_sheet = wb.active
        wb_sheet.cell(3, 1).value = 'Группа'
        wb_sheet.cell(3, 2).value = '№'
        wb_sheet.cell(3, 3).value = 'Фамилия Имя'
        wb_sheet.cell(1, 1).border = top_left_double_border
        wb_sheet.cell(2, 1).border = left_double_border
        wb_sheet.cell(3, 1).border = left_double_border
        wb_sheet.cell(1, 2).border = top_double_border
        wb_sheet.cell(2, 2).border = thin_border
        wb_sheet.cell(3, 2).border = thin_border
        wb_sheet.cell(1, 3).border = top_double_border
        wb_sheet.cell(2, 3).border = thin_border
        wb_sheet.cell(3, 3).border = thin_border

        br = 4
        len_stud_names = []
        stud_id_to_row = {}
        for group_n, studs_ids in studs_id_by_groups.items():
            for i, stud_id, in enumerate(studs_ids):
                stud_info = studs_by_id[stud_id]
                stud_id_to_row[stud_id] = br+i
                stud_name = f'{stud_info.surname} {stud_info.name}'
                len_stud_names.append(len(stud_name))

                wb_sheet.cell(br+i, 1).value = group_n
                wb_sheet.cell(br+i, 1).border = left_double_border
                wb_sheet.cell(br+i, 2).value = stud_info.position_in_group
                wb_sheet.cell(br+i, 2).border = thin_border
                wb_sheet.cell(br+i, 3).value = stud_name
                wb_sheet.cell(br+i, 3).border = right_double_border

            br = wb_sheet.max_row + 1

        fill_header(4, 'Итоговые значения')
        s_id_to_col = {}
        for i, s in enumerate(all_subj):
            s_id_to_col[s.id] = 15+11*i
            fill_header(15+11*i, s.short_title)

        pairs_hierarchy = {}
        for pair in pairs:
            if not pairs_hierarchy.get(pair.group_n):
                pairs_hierarchy[pair.group_n] = {}
            if not pairs_hierarchy[pair.group_n].get(pair.s_id):
                pairs_hierarchy[pair.group_n][pair.s_id] = {}
            if not pairs_hierarchy[pair.group_n][pair.s_id].get(pair.pr_lk_datetime):
                pairs_hierarchy[pair.group_n][pair.s_id][pair.pr_lk_datetime] = []
            if pair.student_id:
                pairs_hierarchy[pair.group_n][pair.s_id][pair.pr_lk_datetime].append(
                    {'stud_id': pair.student_id, 'reason': pair.missing_reason, 'is_debt': pair.is_debt}
                )

        for group_n, subject in pairs_hierarchy.items():
            for s_id, pairs_info in subject.items():
                for pair_dt, missings_info in pairs_info.items():
                    if missings_info:
                        for miss in missings_info:
                            try:
                                reason = miss['reason']
                                studs_id_by_groups[group_n][miss['stud_id']][s_id][reason] += 1
                                if miss['is_debt']:
                                    studs_id_by_groups[group_n][miss['stud_id']][s_id]['долг'] += 1
                            except KeyError:
                                continue

                for stud_id, ss_info in studs_id_by_groups[group_n].items():
                    pairs_count = len(pairs_info)
                    studs_id_by_groups[group_n][stud_id][s_id]['пар'] = pairs_count
                    stud_row = stud_id_to_row[stud_id]
                    s_col = s_id_to_col[s_id]
                    s_info = ss_info[s_id]

                    wb_sheet.cell(stud_row, s_col).value = pairs_count
                    fill_col_with_percent(stud_row, s_col+1, pairs_count, s_info['н'])
                    fill_col_with_percent(stud_row, s_col+3, pairs_count, s_info['б'])
                    fill_col_with_percent(stud_row, s_col+5, pairs_count, s_info['с'])
                    fill_col_with_percent(stud_row, s_col+7, pairs_count, s_info['н']+s_info['б']+s_info['с'])
                    fill_col_with_percent(stud_row, s_col+9, pairs_count, s_info['долг'])

        wb_sheet.column_dimensions['A'].width = 8
        wb_sheet.column_dimensions['B'].width = 5
        wb_sheet.column_dimensions['C'].width = max(len_stud_names) * 1.2
        max_row = wb_sheet.max_row + 1
        for i in range(1, wb_sheet.max_column+1):
            wb_sheet.cell(max_row, i).border = top_only_double_border

        for stud in studs:
            stud_subj = studs_id_by_groups[stud.group_n][stud.id]
            stud_attendance = {'пар': 0, 'н': 0, 'б': 0, 'с': 0, 'долг': 0}
            for attendance in stud_subj.values():
                stud_attendance['пар'] += attendance['пар']
                stud_attendance['н'] += attendance['н']
                stud_attendance['б'] += attendance['б']
                stud_attendance['с'] += attendance['с']
                stud_attendance['долг'] += attendance['долг']

            stud_row = stud_id_to_row[stud.id]
            if stud_attendance['пар'] != 0:
                pairs_count = stud_attendance['пар']
                wb_sheet.cell(stud_row, 4).value = pairs_count
                fill_col_with_percent(stud_row, 5, pairs_count, stud_attendance['н'])
                fill_col_with_percent(stud_row, 7, pairs_count, stud_attendance['б'])
                fill_col_with_percent(stud_row, 9, pairs_count, stud_attendance['с'])
                fill_col_with_percent(stud_row, 11, pairs_count,
                                      stud_attendance['н']+stud_attendance['б']+stud_attendance['с'])
                fill_col_with_percent(stud_row, 13, pairs_count, stud_attendance['долг'])

        faculty = Faculty.objects.values('short_title').get(id=faculty_id)['short_title']
        wb_title = f'Посещаемость {faculty} {course_n} курс{f"{stream_n} поток" if stream_n else ""}.xlsx'
        return {'workbook': wb, 'filename': wb_title}

    else:
        try:
            subj_info = Subject.objects.get(id=subj_id)
        except Subject.DoesNotExist:
            return {'e_msg': 'Какая-то ошибка, не получается обработать запрос'}

        pairs_by_group = {}
        for pair in pairs:
            if not pairs_by_group.get(pair.group_n):
                pairs_by_group[pair.group_n] = {}
            if not pairs_by_group[pair.group_n].get(pair.pr_lk_datetime):
                pairs_by_group[pair.group_n][pair.pr_lk_datetime] = []
            if pair.student_id:
                pairs_by_group[pair.group_n][pair.pr_lk_datetime].append(
                    {'stud_id': pair.student_id, 'reason': pair.missing_reason, 'is_debt': pair.is_debt}
                )

        studs_id_by_groups = {}
        for stud in studs:
            if not studs_id_by_groups.get(stud.group_n):
                studs_id_by_groups[stud.group_n] = {}
            studs_id_by_groups[stud.group_n][stud.id] = {'пар': 0, 'н': 0, 'б': 0, 'с': 0, 'долг': 0}

        wb = Workbook()
        wb_sheet = wb.active
        wb_sheet.cell(3, 1).value = 'Группа'
        wb_sheet.cell(3, 2).value = '№'
        wb_sheet.cell(3, 3).value = 'Фамилия Имя'
        wb_sheet.cell(1, 1).border = top_left_double_border
        wb_sheet.cell(2, 1).border = left_double_border
        wb_sheet.cell(3, 1).border = left_double_border
        wb_sheet.cell(1, 2).border = top_double_border
        wb_sheet.cell(2, 2).border = thin_border
        wb_sheet.cell(3, 2).border = thin_border
        wb_sheet.cell(1, 3).border = top_double_border
        wb_sheet.cell(2, 3).border = thin_border
        wb_sheet.cell(3, 3).border = thin_border

        br = 4
        len_stud_names = []
        stud_id_to_row = {}
        studs_by_id = {stud.id: stud for stud in studs}
        for group_n, studs_ids in studs_id_by_groups.items():
            for i, stud_id, in enumerate(studs_ids):
                stud_info = studs_by_id[stud_id]
                stud_id_to_row[stud_id] = br + i
                stud_name = f'{stud_info.surname} {stud_info.name}'
                len_stud_names.append(len(stud_name))

                wb_sheet.cell(br+i, 1).value = group_n
                wb_sheet.cell(br+i, 1).border = left_double_border
                wb_sheet.cell(br+i, 2).value = stud_info.position_in_group
                wb_sheet.cell(br+i, 2).border = thin_border
                wb_sheet.cell(br+i, 3).value = stud_name
                wb_sheet.cell(br+i, 3).border = right_double_border
            br = wb_sheet.max_row + 1
        wb_sheet.column_dimensions['A'].width = 8
        wb_sheet.column_dimensions['B'].width = 5
        wb_sheet.column_dimensions['C'].width = max(len_stud_names) * 1.2

        fill_header(4, 'Итоговые показатели')

        for i in range(0, subj_info.duration*2, 2):
            wb_sheet.cell(2, 15+i).value = i / 2 + 1
            wb_sheet.cell(2, 15+i).border = left_right_double_border
            wb_sheet.cell(2, 15+i).alignment = center_cell
            wb_sheet.merge_cells(start_row=2, end_row=2, start_column=15+i, end_column=16+i)
            wb_sheet.cell(3, 15+i).value = '+/-'
            wb_sheet.column_dimensions[wb_sheet.cell(3, 15+i).column_letter].width = 5
            wb_sheet.cell(3, 16+i).value = 'отр.'
            wb_sheet.column_dimensions[wb_sheet.cell(3, 16+i).column_letter].width = 6
            for j in range(3, wb_sheet.max_row+1):
                wb_sheet.cell(j, 15+i).border = left_double_border
                wb_sheet.cell(j, 15+i).alignment = center_cell
                wb_sheet.cell(j, 16+i).border = right_double_border
                wb_sheet.cell(j, 16+i).alignment = center_cell
        wb_sheet.cell(1, 15).value = subj_info.short_title
        wb_sheet.cell(1, 15).border = all_double_border
        wb_sheet.cell(1, 15).alignment = center_cell
        wb_sheet.merge_cells(start_row=1, end_row=1, start_column=15, end_column=14+subj_info.duration*2)

        for group_n, pairs_dt in pairs_by_group.items():
            pairs_count = len(pairs_dt)
            for stud_id in studs_id_by_groups[group_n]:
                studs_id_by_groups[group_n][stud_id]['пар'] = pairs_count
            for i, missings in enumerate(pairs_dt.values()):
                for stud_id in studs_id_by_groups[group_n]:
                    wb_sheet.cell(stud_id_to_row[stud_id], 15+i*2).value = '+'
                if missings:
                    for miss in missings:
                        reason = miss['reason']
                        try:
                            stud_row = stud_id_to_row[miss['stud_id']]
                        except KeyError:
                            continue
                        studs_id_by_groups[group_n][miss['stud_id']][reason] += 1
                        if reason == 'н':
                            wb_sheet.cell(stud_row, 15+i*2).value = 'нб'
                            if miss['is_debt']:
                                wb_sheet.cell(stud_row, 16+i*2).value = '-'
                                wb_sheet.cell(stud_row, 16+i*2).font = red_font
                            else:
                                wb_sheet.cell(stud_row, 16+i*2).value = '+'
                        elif reason == 'б':
                            wb_sheet.cell(stud_row, 15+i*2).value = 'бол'
                            if miss['is_debt']:
                                wb_sheet.cell(stud_row, 16+i*2).value = '-'
                                wb_sheet.cell(stud_row, 16+i*2).font = red_font
                            else:
                                wb_sheet.cell(stud_row, 16+i*2).value = '+'
                        elif reason == 'c':
                            wb_sheet.cell(stud_row, 15+i*2).value = 'спр'
            for stud_id in studs_id_by_groups[group_n]:
                stud_row = stud_id_to_row[stud_id]
                stud_attendance = studs_id_by_groups[group_n][stud_id]
                if stud_attendance['пар'] != 0:
                    pairs_count = stud_attendance['пар']
                    wb_sheet.cell(stud_row, 4).value = pairs_count
                    fill_col_with_percent(stud_row, 5, pairs_count, stud_attendance['н'])
                    fill_col_with_percent(stud_row, 7, pairs_count, stud_attendance['б'])
                    fill_col_with_percent(stud_row, 9, pairs_count, stud_attendance['с'])
                    fill_col_with_percent(stud_row, 11, pairs_count,
                                          stud_attendance['н']+stud_attendance['б']+stud_attendance['с'])
                    fill_col_with_percent(stud_row, 13, pairs_count, stud_attendance['долг'])

        faculty = Faculty.objects.values('short_title').get(id=faculty_id)['short_title']
        wb_title = f'Посещаемость {faculty} {course_n} курс {f"{stream_n} поток" if stream_n else ""}' \
                   f' {subj_info.short_title}.xlsx'
        return {'workbook': wb, 'filename': wb_title}


def show_attendance_hours_course(faculty_id, course_n, stream_n=None, gr=None, fill_subjects=True):
    sql_request = '''SELECT
            ps.id,
            ps.group_n,
            ps.pr_lk_datetime,
            s.id AS s_id,
            s.short_title,
            s.duration,
            ps.is_practical AS is_pr,
            mp.student_id,
            mp.missing_reason,
            mp.is_debt
        FROM reports_pairsschedule AS ps
            JOIN reports_subject AS s
            ON s.id = ps.subject_id
            LEFT JOIN reports_missingpair AS mp
            ON mp.pair_id = ps.id
        WHERE
            ps.faculty_id = %s AND
            ps.course_n = %s AND
            {}
            ps.is_reported = %s
        ORDER BY
            ps.group_n,
            s.id,
            ps.is_practical,
            ps.pr_lk_datetime
            '''
    sql_pars = [faculty_id, course_n, True]

    if stream_n:
        studs = tuple(Student.objects.filter(faculty=faculty_id, course_n=course_n, stream_n=stream_n, is_fired=False))
        groups = []
        for stud in studs:
            if stud.group_n not in groups:
                groups.append(stud.group_n)
        sql_request = sql_request.format('ps.group_n = ANY(%s) AND')
        sql_pars.insert(2, groups)
    elif gr:
        sql_request = sql_request.format('ps.group_n = %s AND')
        sql_pars.insert(2, gr)
        studs = tuple(Student.objects.filter(faculty=faculty_id, course_n=course_n, group_n=gr, is_fired=False))
    else:
        sql_request = sql_request.format('')
        studs = tuple(Student.objects.filter(faculty=faculty_id, course_n=course_n, is_fired=False))

    pairs = tuple(PairsSchedule.objects.raw(sql_request, sql_pars))
    if not pairs:
        return {'e_msg': 'Ни одной рапортички ещё не было'}
    all_subjects = tuple(Subject.objects.filter(faculty_id=faculty_id, course_n=course_n))
    subj_by_id = {s.id: s for s in all_subjects}

    studs_id_by_group = {}
    for stud in studs:
        if not studs_id_by_group.get(stud.group_n):
            studs_id_by_group[stud.group_n] = {}
        studs_id_by_group[stud.group_n][stud.id] = {
            subj.id: {
                True: {'н': 0, 'б': 0, 'с': 0, 'долг': 0},
                False: {'н': 0, 'б': 0, 'с': 0, 'долг': 0}
            } for subj in all_subjects
        }
        studs_id_by_group[stud.group_n][stud.id]['пр'] = {'часов': 0, 'н': 0, 'б': 0, 'с': 0, 'долг': 0}
        studs_id_by_group[stud.group_n][stud.id]['лк'] = {'часов': 0, 'н': 0, 'б': 0, 'с': 0, 'долг': 0}

    pairs_hierarchy = {}
    for pair in pairs:
        if not pairs_hierarchy.get(pair.group_n):
            pairs_hierarchy[pair.group_n] = {}
        if not pairs_hierarchy[pair.group_n].get(pair.s_id):
            pairs_hierarchy[pair.group_n][pair.s_id] = {}
        if not pairs_hierarchy[pair.group_n][pair.s_id].get(pair.is_practical):
            pairs_hierarchy[pair.group_n][pair.s_id][pair.is_practical] = {}
        if not pairs_hierarchy[pair.group_n][pair.s_id][pair.is_practical].get(pair.pr_lk_datetime):
            pairs_hierarchy[pair.group_n][pair.s_id][pair.is_practical][pair.pr_lk_datetime] = []
        if pair.student_id:
            pairs_hierarchy[pair.group_n][pair.s_id][pair.is_practical][pair.pr_lk_datetime].append(
                {'stud_id': pair.student_id, 'reason': pair.missing_reason, 'is_debt': pair.is_debt}
            )

    wb = Workbook()
    wb_sheet = wb.active
    wb_sheet.cell(3, 1).value = 'Группа'
    wb_sheet.cell(3, 2).value = '№'
    wb_sheet.cell(3, 3).value = 'Фамилия Имя'
    wb_sheet.cell(1, 1).border = top_left_double_border
    wb_sheet.cell(2, 1).border = left_double_border
    wb_sheet.cell(3, 1).border = left_double_border
    wb_sheet.cell(1, 2).border = top_double_border
    wb_sheet.cell(2, 2).border = thin_border
    wb_sheet.cell(3, 2).border = thin_border
    wb_sheet.cell(1, 3).border = top_double_border
    wb_sheet.cell(2, 3).border = thin_border
    wb_sheet.cell(3, 3).border = thin_border

    br = 4
    len_stud_names = []
    stud_id_to_row = {}
    studs_by_id = {stud.id: stud for stud in studs}
    for group_n, studs_ids in studs_id_by_group.items():
        for i, stud_id, in enumerate(studs_ids):
            stud_info = studs_by_id[stud_id]
            stud_id_to_row[stud_id] = br + i
            stud_name = f'{stud_info.surname} {stud_info.name}'
            len_stud_names.append(len(stud_name))

            wb_sheet.cell(br+i, 1).value = group_n
            wb_sheet.cell(br+i, 1).border = left_double_border
            wb_sheet.cell(br+i, 2).value = stud_info.position_in_group
            wb_sheet.cell(br+i, 2).border = thin_border
            wb_sheet.cell(br+i, 3).value = stud_name
            wb_sheet.cell(br+i, 3).border = right_double_border

        br = wb_sheet.max_row + 1

    wb_sheet.column_dimensions['A'].width = 8
    wb_sheet.column_dimensions['B'].width = 5
    wb_sheet.column_dimensions['C'].width = max(len_stud_names) * 1.2

    def fill_header(bc, title):
        wb_sheet.cell(1, bc).value = title
        wb_sheet.cell(1, bc).border = all_double_border
        wb_sheet.cell(1, bc).alignment = center_cell
        wb_sheet.merge_cells(start_row=1, end_row=1, start_column=bc, end_column=bc+10)
        wb_sheet.cell(2, bc).value = 'часов'
        wb_sheet.cell(2, bc).border = left_double_border
        wb_sheet.cell(2, bc).alignment = center_cell
        wb_sheet.cell(3, bc).value = 'ч.'
        wb_sheet.cell(2, bc+1).value = 'причина нб'
        wb_sheet.cell(2, bc+1).border = thin_border
        wb_sheet.cell(2, bc+1).alignment = center_cell
        wb_sheet.merge_cells(start_row=2, end_row=2, start_column=bc+1, end_column=bc+6)
        wb_sheet.cell(3, bc+1).value = 'н,ч.'
        wb_sheet.cell(3, bc+2).value = 'н,%'
        wb_sheet.cell(3, bc+3).value = 'б,ч.'
        wb_sheet.cell(3, bc+4).value = 'б,%'
        wb_sheet.cell(3, bc+5).value = 'с,ч.'
        wb_sheet.cell(3, bc+6).value = 'с,%'
        wb_sheet.cell(2, bc+7).value = 'всего нб'
        wb_sheet.cell(2, bc+7).border = thin_border
        wb_sheet.cell(2, bc+7).alignment = center_cell
        wb_sheet.merge_cells(start_row=2, end_row=2, start_column=bc+7, end_column=bc+8)
        wb_sheet.cell(3, bc+7).value = 'ч.'
        wb_sheet.cell(3, bc+8).value = '%'
        wb_sheet.cell(2, bc+9).value = 'неотраб.'
        wb_sheet.cell(2, bc+9).border = right_double_border
        wb_sheet.cell(2, bc+9).alignment = center_cell
        wb_sheet.merge_cells(start_row=2, end_row=2, start_column=bc+9, end_column=bc+10)
        wb_sheet.cell(3, bc+9).value = 'ч.'
        wb_sheet.cell(3, bc+10).value = '%'

        for ii in range(3, wb_sheet.max_row+1):
            wb_sheet.cell(ii, bc).border = left_double_border
            wb_sheet.cell(ii, bc).alignment = center_cell
            wb_sheet.cell(ii, bc+10).border = right_double_border
            wb_sheet.cell(ii, bc+10).alignment = center_cell
            for col in range(1, 10):
                wb_sheet.cell(ii, bc+col).border = thin_border
                wb_sheet.cell(ii, bc+col).alignment = center_cell
        for ii in range(bc, bc+11):
            if (ii - bc) % 2 == 0:
                wb_sheet.column_dimensions[wb_sheet.cell(3, ii).column_letter].width = 5.5
            else:
                wb_sheet.column_dimensions[wb_sheet.cell(3, ii).column_letter].width = 4.5

    fill_header(4, 'Итоговые значения')
    fill_header(15, 'Практические занятия')
    fill_header(26, 'Лекции')

    s_id_to_col = {}
    for i, s in enumerate(all_subjects):
        s_id_to_col[s.id] = {True: 37 + 22 * i, False: 48 + 22 * i}
        if fill_subjects:
            fill_header(
                37 + 22 * i,
                f'{s.short_title} - практические занятия ({s.hours_per_pair}ч., всего {s.hours_per_pair*s.duration}ч.)'
            )
            fill_header(
                48 + 22 * i,
                f'{s.short_title} - лекции (2ч.)'
            )

    def fill_col_with_percent(st_row, subj_col, p_count, n_count):
        wb_sheet.cell(st_row, subj_col).value = n_count
        percent_n = round(100 * n_count / p_count, 1) if p_count != 0 else 0
        wb_sheet.cell(st_row, subj_col + 1).value = percent_n
        if percent_n > 30:
            wb_sheet.cell(st_row, subj_col + 1).font = red_font

    for group_n, subject in pairs_hierarchy.items():
        for s_id, pairs_type in subject.items():
            for is_pr, pairs_info in pairs_type.items():
                for pair_dt, missings_info in pairs_info.items():
                    if missings_info:
                        for miss in missings_info:
                            try:
                                reason = miss['reason']
                                studs_id_by_group[group_n][miss['stud_id']][s_id][is_pr][reason] += 1
                                if miss['is_debt']:
                                    studs_id_by_group[group_n][miss['stud_id']][s_id][is_pr]['долг'] += 1
                            except KeyError:
                                continue

                pair_hours = subj_by_id[s_id].hours_per_pair if is_pr else 2
                hours_count = len(pairs_info) * pair_hours
                s_col = s_id_to_col[s_id][is_pr]
                for stud_id, ss_info in studs_id_by_group[group_n].items():
                    s_info = ss_info[s_id][is_pr]
                    is_pr_key = 'пр' if is_pr else 'лк'
                    this_stud_pr_lk = studs_id_by_group[group_n][stud_id][is_pr_key]
                    this_stud_pr_lk['часов'] += hours_count

                    n_hours = s_info['н'] * pair_hours
                    this_stud_pr_lk['н'] += n_hours

                    b_hours = s_info['б'] * pair_hours
                    this_stud_pr_lk['б'] += b_hours

                    c_hours = s_info['с'] * pair_hours
                    this_stud_pr_lk['с'] += c_hours

                    debt_hours = s_info['долг'] * pair_hours
                    this_stud_pr_lk['долг'] += debt_hours

                    if fill_subjects:
                        stud_row = stud_id_to_row[stud_id]
                        wb_sheet.cell(stud_row, s_col).value = hours_count
                        fill_col_with_percent(stud_row, s_col + 1, hours_count, n_hours)
                        fill_col_with_percent(stud_row, s_col + 3, hours_count, b_hours)
                        fill_col_with_percent(stud_row, s_col + 5, hours_count, c_hours)
                        fill_col_with_percent(stud_row, s_col + 7, hours_count, n_hours + b_hours + c_hours)
                        fill_col_with_percent(stud_row, s_col + 9, hours_count, debt_hours)

    for group_n, group_studs in studs_id_by_group.items():
        for stud_id, attendance in group_studs.items():
            stud_row = stud_id_to_row[stud_id]
            pr = attendance['пр']
            lk = attendance['лк']
            sum_hours = pr['часов'] + lk['часов']
            if sum_hours != 0:
                wb_sheet.cell(stud_row, 4).value = sum_hours
                wb_sheet.cell(stud_row, 15).value = pr['часов']
                wb_sheet.cell(stud_row, 26).value = lk['часов']

                fill_col_with_percent(stud_row, 5, sum_hours, pr['н'] + lk['н'])
                fill_col_with_percent(stud_row, 16, pr['часов'], pr['н'])
                fill_col_with_percent(stud_row, 27, lk['часов'], lk['н'])

                fill_col_with_percent(stud_row, 7, sum_hours, pr['б'] + lk['б'])
                fill_col_with_percent(stud_row, 18, pr['часов'], pr['б'])
                fill_col_with_percent(stud_row, 29, lk['часов'], lk['б'])

                fill_col_with_percent(stud_row, 9, sum_hours, pr['с'] + lk['с'])
                fill_col_with_percent(stud_row, 20, pr['часов'], pr['с'])
                fill_col_with_percent(stud_row, 31, lk['часов'], lk['с'])

                fill_col_with_percent(stud_row, 11, sum_hours, pr['н'] + lk['н'] + pr['б'] + lk['б'] + pr['с'] + lk['с'])
                fill_col_with_percent(stud_row, 22, pr['часов'], pr['н'] + pr['б'] + pr['с'])
                fill_col_with_percent(stud_row, 33, lk['часов'], lk['н'] + lk['б'] + lk['с'])

                fill_col_with_percent(stud_row, 13, sum_hours, pr['долг'] + lk['долг'])
                fill_col_with_percent(stud_row, 24, pr['часов'], pr['долг'])
                fill_col_with_percent(stud_row, 35, lk['часов'], lk['долг'])

    faculty = Faculty.objects.values('short_title').get(id=faculty_id)['short_title']
    group_or_stream = f' {stream_n} поток' if stream_n else f' {gr} группа' if gr else ''
    wb_title = f'Посещаемость по часам {faculty} {course_n} курс{group_or_stream}.xlsx'
    return {'workbook': wb, 'filename': wb_title}
