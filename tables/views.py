from uuid import UUID

from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, FormView, DetailView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import save_virtual_workbook
from transliterate import translit

from reports.models import Student
from vk_bot.vk_funcs.config import var_message
from .models import Table, TablePart
from .forms import CreateTableForm, FillTableForm


class TablesListView(LoginRequiredMixin, ListView):
    model = Table
    context_object_name = 'tables'
    template_name = 'tables/list.html'
    extra_context = {'title': 'Таблицы'}

    def get_queryset(self):
        return self.model.objects.filter(owner=self.request.user)


class NewTable(LoginRequiredMixin, FormView):
    form_class = CreateTableForm
    template_name = 'tables/create.html'
    extra_context = {'title': 'Новая таблица'}

    def get_success_url(self):
        return reverse_lazy('tables_show', args=(self.object.id,))

    def form_valid(self, form):
        self.object = form.save()
        return super().form_valid(form)

    def get_form(self, form_class=None):
        form = super().get_form(form_class=form_class)
        user = self.request.user
        if not user.is_superuser:
            form.fields['faculty'].disabled = True
        if self.request.method == 'POST':
            form.instance.owner = self.request.user
        return form


class ShowTable(LoginRequiredMixin, DetailView):
    model = Table
    queryset = Table.objects.prefetch_related('parts').all()
    context_object_name = 'table'
    template_name = 'tables/show.html'

    def get_queryset(self):
        return self.queryset.filter(owner_id=self.request.user.id)


class UpdateTable(LoginRequiredMixin, UpdateView):
    model = Table
    queryset = Table.objects.prefetch_related('parts').all()
    template_name = 'tables/edit.html'
    context_object_name = 'table'
    fields = ('title', 'description')

    def get_success_url(self):
        return reverse('tables_show', args=(self.object.id,))

    def get_queryset(self):
        return self.queryset.filter(owner_id=self.request.user.id)


class FillTable(UpdateView):
    model = TablePart
    queryset = TablePart.objects.select_related('document').all()
    form_class = FillTableForm
    context_object_name = 'part'
    template_name = 'tables/fill.html'

    def get_form_kwargs(self):
        print(self.object, id(self))
        return super().get_form_kwargs()

    def get_success_url(self):
        print(self.object, id(self))
        return reverse('tables_fill_success', args=(self.object.id,))


def fill_success(request, pk):
    return render(request, 'tables/fill_success.html', context={"object_id": pk})


class ExportExcelView(LoginRequiredMixin, DetailView):
    model = Table
    object: Table
    queryset = Table.objects.prefetch_related('parts').all()

    def get_queryset(self):
        return self.queryset.filter(owner_id=self.request.user.id)

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        excel = self.create_excel()
        response = HttpResponse(save_virtual_workbook(excel), content_type='application/ms-excel')
        filename = translit(self.object.title.replace(' ', '_').replace(',', '_'), 'ru', True)
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        return response

    def create_excel(self):
        table = self.object
        wb = Workbook()
        sheet: Worksheet = wb.active
        sheet.cell(1, 1).value = 'Группа'
        sheet.cell(1, 2).value = 'ФИО'
        for c, field in enumerate(table.fields, start=3):
            sheet.cell(1, c).value = field["name"]
        r = 2
        for part in table.parts.all():
            for row in part.data:
                sheet.cell(r, 1).value = part.group_n
                sheet.cell(r, 2).value = row['fio']
                for c, field in enumerate(table.fields, start=3):
                    sheet.cell(r, c).value = row[field['alias']]
                r += 1
        return wb


class NotifyView(ShowTable):
    def get(self, request, *args, **kwargs):
        self.object = table = self.get_object()
        heads = Student.objects. \
            filter(faculty=table.faculty, course_n=table.course_n, is_head=True, vk_id__isnull=False)
        heads = {h.group_n: h for h in heads}
        for part in table.parts.all():
            head = heads.get(part.group_n)
            if not head or not head.vk_id:
                continue
            try:
                var_message(head.vk_id, text=self.table_message(part.id))
            except:
                pass
        return redirect(reverse('tables_show', args=(self.object.id,)))

    def table_message(self, part_uuid: UUID) -> str:
        return f'Нужно заполнить таблицу: {self.object.title}\n' \
               f'{settings.SITE_HOST}{reverse("tables_fill", args=(part_uuid,))}\n' \
               f'{self.object.description}'
